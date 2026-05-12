"""
Part 2 of UAT 테스트북 v2 builder.
Provides: 성과, 채용, 보상, 인사이트·분석, 온보딩·오프보딩, 규정준수, My Space, V2 홈, 알림, 엣지케이스, 집계
Called from build-testbook-v2.py via attach_part2(wb, scenario_counts).
"""

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Import shared styles + helpers from main module
import sys
import importlib.util
from pathlib import Path

_main_path = Path(__file__).parent / 'build-testbook-v2.py'
_spec = importlib.util.spec_from_file_location('build_testbook_v2_main', _main_path)
_main = importlib.util.module_from_spec(_spec)
sys.modules['build_testbook_v2_main'] = _main
_spec.loader.exec_module(_main)

# Re-export
SCENARIO_HEADERS = _main.SCENARIO_HEADERS
SCENARIO_WIDTHS = _main.SCENARIO_WIDTHS
CHECKLIST_HEADERS = _main.CHECKLIST_HEADERS
CHECKLIST_WIDTHS = _main.CHECKLIST_WIDTHS
apply_header = _main.apply_header
write_scenario_rows = _main.write_scenario_rows
write_checklist_rows = _main.write_checklist_rows
add_dropdown = _main.add_dropdown
add_conditional_pass_fail = _main.add_conditional_pass_fail
SCENARIO_TAB = _main.SCENARIO_TAB
CHECKLIST_TAB = _main.CHECKLIST_TAB
META_TAB = _main.META_TAB
EDGE_TAB = _main.EDGE_TAB
FONT_NAME = _main.FONT_NAME
HEADER_FONT = _main.HEADER_FONT
HEADER_FILL = _main.HEADER_FILL
BODY_FONT = _main.BODY_FONT
TITLE_FONT = _main.TITLE_FONT
SECTION_FONT = _main.SECTION_FONT
ALIGN_TOP_WRAP = _main.ALIGN_TOP_WRAP
ALIGN_CENTER = _main.ALIGN_CENTER
BORDER = _main.BORDER
NEW_FILL = _main.NEW_FILL


def sheet_performance(wb):
    """성과 — scenario sheet with new GoalRevision/QuarterlyReview/Calibration DnD/AI eval"""
    ws = wb.create_sheet('성과')
    ws.sheet_properties.tabColor = SCENARIO_TAB
    apply_header(ws, SCENARIO_HEADERS, SCENARIO_WIDTHS)

    rows = [
        (1, '성과 대시보드 진입', 'HR_ADMIN', '로그인 완료', '',
         '좌측 메뉴 > 성과/보상 > 성과 관리 클릭', '성과 대시보드가 표시됨 (MBO/CFR 통계)'),
        (2, '새 목표 생성 (EMPLOYEE)', 'EMPLOYEE (employee-a@ctr.co.kr)',
         '현재 평가 사이클 진행 중', '목표 제목/설명/측정 지표/기한',
         '나의 공간 > 목표/평가 > 새 목표 생성 > 입력 > 저장',
         '목표가 생성되고 진행률 0%로 표시됨'),
        (3, 'GoalRevision 배치 수정', 'EMPLOYEE / MANAGER', 'Step 2 완료',
         '진행 중인 목표를 분기 중 수정',
         '목표 상세 > 수정 > 변경 사유 입력 > 제출 → 매니저 승인',
         '목표 버전(GoalRevision)이 생성되고 이전 버전 이력 보존. QGP 분기 데이터 오염 방어 (기존 데이터는 이전 버전에 귀속)'),
        (4, '자기평가 작성', 'EMPLOYEE (employee-a@ctr.co.kr)', '평가 사이클 자기평가 단계', '',
         '/performance/self-eval 진입 > 평가 항목별 점수 + 코멘트 입력 > 제출',
         '자기평가가 제출되고 매니저 평가 단계로 전환됨'),
        (5, 'AI 자기평가 도우미', 'EMPLOYEE (employee-a@ctr.co.kr)', '자기평가 작성 중', '',
         '자기평가 작성 화면에서 "AI 코멘트 초안 생성" 버튼 클릭',
         'Claude 기반 자기평가 코멘트 초안이 생성되고 직원이 편집하여 제출 가능'),
        (6, '매니저 평가 작성', 'MANAGER (manager@ctr.co.kr)', 'Step 4 완료', '',
         '/performance/manager-eval 또는 /performance/manager-evaluation 진입 (i18n 신버전)',
         '매니저가 직원 평가를 작성하고 제출 가능. 두 라우트 동시 존재 시 사용성 확인'),
        (7, '동료평가 설정', 'HR_ADMIN', '평가 사이클 동료평가 단계', '',
         '/performance/peer-review/[cycleId]/setup 진입 > 평가자/피평가자 매칭 > 저장',
         '동료평가 매칭이 설정되고 평가자에게 알림 발송'),
        (8, '동료평가 실행', 'EMPLOYEE / MANAGER', 'Step 7 완료', '',
         '/performance/peer-review/evaluate/[id] 진입 > 평가 작성 > 제출',
         '동료평가가 제출됨 (5-record privacy 적용)'),
        (9, '분기 리뷰 — 직원 제출', 'EMPLOYEE (employee-a@ctr.co.kr)', '분기 리뷰 단계 진행 중', '',
         '나의 공간 > 분기 리뷰 > 작성 > 제출',
         '직원 분기 리뷰가 제출됨'),
        (10, '분기 리뷰 — 매니저 듀얼 제출', 'MANAGER (manager@ctr.co.kr)', 'Step 9 완료', '',
         '팀 관리 > 분기 리뷰 > 해당 팀원 선택 > 작성 > 제출',
         '매니저 분기 리뷰가 별도로 제출됨 (직원/매니저 듀얼 제출 + 마스킹 정책 적용)'),
        (11, '분기 리뷰 마스킹 확인', 'EMPLOYEE (employee-a@ctr.co.kr)', 'Step 10 완료', '',
         '나의 공간 > 분기 리뷰 상세 진입',
         '매니저 코멘트가 정해진 시점까지 마스킹되어 표시됨'),
        (12, '캘리브레이션 DnD 배치 조정', 'HR_ADMIN', '캘리브레이션 단계 진행 중', '',
         '/performance/calibration 진입 > 직원 카드 드래그 > 9-block 셀로 드롭 (다중선택 지원)',
         '@dnd-kit 기반 드래그·드롭으로 등급 조정 완료 + 배치 API로 일괄 저장'),
        (13, '보상 검토 (Merit Matrix)', 'HR_ADMIN', '캘리브레이션 완료', '',
         '/performance/comp-review 진입 > 성과등급별 merit 매트릭스 확인',
         '직원별 성과등급 + 인상률 + 성과급 시뮬레이션이 매트릭스로 표시됨'),
        (14, '성장 여정 시각화 (MyResult)', 'EMPLOYEE (employee-a@ctr.co.kr)',
         '과거 평가 사이클 완료 이력 존재', '',
         '/performance/my-result 진입',
         '과거 사이클 추이가 라인 차트로 시각화됨 (성장 여정)'),
        (15, '1:1 미팅 설정', 'MANAGER (manager@ctr.co.kr)', '로그인 완료',
         '1:1 대상: 이민준, 주기: 격주',
         '팀 관리 > 1:1 미팅 > 새 미팅 설정 > 저장',
         '1:1 미팅 일정이 생성되고 양측에 알림'),
        (16, '인정/칭찬 피드', 'EMPLOYEE (employee-a@ctr.co.kr)', '로그인 완료', '',
         '나의 공간 > 리코그니션 진입 > 동료 인정 작성 > 제출',
         '인정 메시지가 피드에 표시됨'),
        (17, '펄스 설문 응답', 'EMPLOYEE (employee-a@ctr.co.kr)', '펄스 설문 활성', '',
         '/performance/pulse/[id]/respond 진입 > 응답 > 제출',
         '펄스 응답이 제출되고 결과는 익명 집계됨'),
        (18, '펄스 결과 조회', 'HR_ADMIN', 'Step 17 완료', '',
         '/performance/pulse/[id]/results 진입',
         '응답 결과가 카테고리별/팀별로 집계되어 표시됨'),
        (19, '성과 알림', 'EMPLOYEE / MANAGER', '평가 마감일 D-day 경과', '',
         '/performance/notifications 진입 또는 알림 센터',
         '평가 마감 임박/경과 알림이 발송됨 (eval-reminder cron)'),
        (20, '바이어스 감지', 'HR_ADMIN', 'AI 평가 초안 사용', '',
         '매니저 평가에서 AI 초안 생성 → 바이어스 감지 결과 확인',
         '성별/연령별/근속 등 잠재적 바이어스 패턴이 감지되어 경고됨'),
    ]
    write_scenario_rows(ws, rows, mark_new={3, 5, 10, 11, 12, 14, 20})
    add_dropdown(ws, f'I2:I{len(rows) + 1}', ['Pass', 'Fail', 'Block', 'N/A'])
    add_conditional_pass_fail(ws, f'I2:I{len(rows) + 1}')
    return len(rows)


def sheet_recruitment(wb):
    """채용 — checklist"""
    ws = wb.create_sheet('채용')
    ws.sheet_properties.tabColor = CHECKLIST_TAB
    apply_header(ws, CHECKLIST_HEADERS, CHECKLIST_WIDTHS)

    rows = [
        (1, '채용', '채용 공고 목록 (/recruitment)', 'HR_ADMIN'),
        (2, '채용', '새 공고 작성 (/recruitment/new)', 'HR_ADMIN'),
        (3, '채용', '공고 상세 (/recruitment/[id])', 'HR_ADMIN'),
        (4, '채용', '공고 수정 (/recruitment/[id]/edit)', 'HR_ADMIN'),
        (5, '지원자', '지원자 목록 (/recruitment/[id]/applicants)', 'HR_ADMIN'),
        (6, '지원자', '새 지원자 등록 (/recruitment/[id]/applicants/new)', 'HR_ADMIN'),
        (7, '면접', '면접 일정 목록 (/recruitment/[id]/interviews)', 'HR_ADMIN'),
        (8, '면접', '새 면접 등록 + 타임존 자동 설정 (/recruitment/[id]/interviews/new)', 'HR_ADMIN'),
        (9, '면접', '면접 중복 감지', 'HR_ADMIN'),
        (10, '파이프라인', '10-stage 칸반 파이프라인 (/recruitment/[id]/pipeline)', 'HR_ADMIN'),
        (11, '파이프라인', 'OFFER_ACCEPTED / OFFER_DECLINED 단계 표시', 'HR_ADMIN'),
        (12, '보드 뷰', '전체 보드 뷰 (/recruitment/board)', 'HR_ADMIN'),
        (13, '대시보드', '채용 대시보드 (/recruitment/dashboard)', 'HR_ADMIN'),
        (14, '대시보드', '채용 퍼널에 OFFER_ACCEPTED/DECLINED 단계 시각화', 'HR_ADMIN'),
        (15, '인재풀', '인재풀 관리 (/recruitment/talent-pool)', 'HR_ADMIN'),
        (16, '비용 분석', '채용 비용 분석 (/recruitment/cost-analysis)', 'HR_ADMIN'),
        (17, '채용요청서', '채용요청서 목록 (/recruitment/requisitions)', 'HR_ADMIN'),
        (18, '채용요청서', '새 채용요청서 (/recruitment/requisitions/new)', 'HR_ADMIN'),
        (19, '채용요청서', '채용요청서 상세 + canApprove 서버 결정 (PR #18)', 'multi-role 사용자'),
        (20, '내부공모', '내부 공모 확인 (/my/internal-jobs)', 'EMPLOYEE'),
        (21, 'ATS 전환', 'ATS 합격자 → 직원 원클릭 전환', 'HR_ADMIN'),
        (22, '오퍼 Flow', '오퍼 수락/거절 상태머신 + PATCH API + 다이얼로그', 'HR_ADMIN'),
        (23, '도메인 이벤트', 'OFFER_SENT/ACCEPTED/DECLINED + INTERVIEW_SCHEDULED 알림', 'HR_ADMIN'),
        (24, 'do-not-rehire', '재고용 방지 직원의 채용 중복 체크 경고', 'HR_ADMIN'),
        (25, 'AI 스크리닝', 'AI 기반 지원자 스크리닝 결과', 'HR_ADMIN'),
    ]
    write_checklist_rows(ws, rows)
    add_dropdown(ws, f'E2:E{len(rows) + 1}', ['Y', 'N', 'NA'])
    add_dropdown(ws, f'F2:F{len(rows) + 1}', ['Critical', 'Major', 'Minor', '개선'])
    return len(rows)


def sheet_compensation(wb):
    """보상 — EXPANDED scenario with Off-Cycle 3 paths + Comp Letter + Total Rewards"""
    ws = wb.create_sheet('보상')
    ws.sheet_properties.tabColor = SCENARIO_TAB
    apply_header(ws, SCENARIO_HEADERS, SCENARIO_WIDTHS)

    rows = [
        (1, '보상 대시보드 (HR)', 'HR_ADMIN', '로그인 완료', '',
         '좌측 메뉴 > 성과/보상 > 보상 관리 클릭',
         '보상 대시보드가 표시되고 정기/비정기 조정 현황 보임'),
        (2, '정기 보상 검토 (merit matrix)', 'HR_ADMIN', '캘리브레이션 완료', '',
         '/performance/comp-review 진입 > 인상률 매트릭스 확인',
         '직급별 인상률 + PayBandChart로 밴드 내 위치 시각화'),
        (3, 'Off-Cycle — 매니저 제안 경로', 'MANAGER (manager@ctr.co.kr)',
         '팀원 대상', '대상: 이민준, 인상률: 5%, 시행일: 2026-06-01',
         '팀 관리 > 비정기 조정 제안 > 대상/사유/금액 입력 > 제출',
         '매니저 제안이 HR 결재함으로 전달되고 ApprovalFlow 다단계 진행'),
        (4, 'Off-Cycle — HR 주도 경로', 'HR_ADMIN', '로그인 완료',
         '대상: 정다은, 사유: 시장 보정',
         '/compensation/off-cycle/new > HR 주도 발의 옵션 선택 > 저장',
         'HR 발의 건이 결재 라인 진행 (HR_ADMIN 발의 시 step 2 self-skip — PR #19)'),
        (5, 'Off-Cycle — 자동 트리거', 'HR_ADMIN', '발령(승진 등) 처리 직후', '',
         '발령 처리 → 자동으로 비정기 조정 결재 생성됨을 확인',
         '발령 시 자동 트리거된 비정기 조정 건이 결재 대기 목록에 표시됨'),
        (6, '급여 역전 방어 검증', 'HR_ADMIN', '부하 연봉 > 상사 시나리오 데이터', '',
         '/compensation/off-cycle/new > 부하 직원 대상으로 상사 연봉 초과하는 인상 시도',
         '저장 전 경고 표시 (급여 역전 방지 정책)'),
        (7, 'Self-approval Skip', 'HR_ADMIN (발의자 = 결재자)', '본인 발의 → 본인 결재 단계 포함',
         '발의자가 동시에 step의 결재자인 경우',
         '결재 단계 진행',
         '발의자 본인이 결재자인 step은 자동 self-skip 처리 (APPROVED)'),
        (8, '미래 시행일 Cron', 'HR_ADMIN', '시행일 > 오늘', '',
         '/compensation/off-cycle/new > 시행일을 미래로 설정 > 결재 완료',
         '결재는 완료되되, 실제 급여 반영은 시행일 cron이 처리 (즉시 반영 X)'),
        (9, 'Compensation Letter PDF 생성', 'HR_ADMIN', '결재 완료된 보상 조정 건 존재', '',
         '보상 상세 > "Compensation Letter 생성" 클릭',
         'PDF가 생성되고 S3에 저장됨 (compensation-letter-pdf.ts)'),
        (10, 'Compensation Letter 이메일 발송 (배치)', 'HR_ADMIN', 'Step 9 완료',
         '복수 직원 대상 일괄 발송',
         '보상 관리 > 레터 배치 발송 > 대상 선택 > 발송',
         '이메일이 일괄 발송되고 발송 이력이 기록됨 (/api/v1/compensation/letters/send)'),
        (11, 'Compensation Letter 버전 관리', 'HR_ADMIN', '한 직원에 복수 레터 발송 이력', '',
         '직원 상세 > 보상 레터 이력 확인',
         '동일 직원의 여러 레터가 버전 번호와 함께 표시됨'),
        (12, 'Total Rewards Statement (직원)', 'EMPLOYEE (employee-a@ctr.co.kr)', '로그인 완료', '',
         '나의 공간 > 내 총보상 (/my/total-rewards)',
         '기본급/상여/수당/복리후생/포상 연간 집계 + 도넛 차트 + 트렌드 라인이 표시됨'),
        (13, 'PayBandChart 표시 — Profile', 'EMPLOYEE (employee-a@ctr.co.kr)', '로그인 완료', '',
         '내 프로필 > Compensation 영역 확인',
         'PayBandChart compact 버전이 ProfileSidebar에 표시됨'),
        (14, 'PayBandChart 표시 — Simulation', 'HR_ADMIN', '보상 시뮬레이션', '',
         '/compensation/off-cycle/new > 시뮬레이션 탭',
         'before/after PayBandChart로 변경 전후 비교'),
        (15, '직급별 보상 분포 분석', 'HR_ADMIN', '로그인 완료', '',
         '/analytics/compensation 진입',
         '직급별 연봉 분포 + 성별 임금격차 (GenderPayGap) 분석 표시'),
    ]
    write_scenario_rows(ws, rows, mark_new={3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14})
    add_dropdown(ws, f'I2:I{len(rows) + 1}', ['Pass', 'Fail', 'Block', 'N/A'])
    add_conditional_pass_fail(ws, f'I2:I{len(rows) + 1}')
    return len(rows)


def sheet_insights(wb):
    """인사이트·분석 — checklist"""
    ws = wb.create_sheet('인사이트·분석')
    ws.sheet_properties.tabColor = CHECKLIST_TAB
    apply_header(ws, CHECKLIST_HEADERS, CHECKLIST_WIDTHS)

    rows = [
        (1, '분석', '경영진 요약 대시보드 (/analytics)', 'EXECUTIVE / HR_ADMIN'),
        (2, '인력 분석', '인력 현황 분석 (/analytics/workforce)', 'HR_ADMIN'),
        (3, '급여 분석', '급여 통계 (/analytics/payroll)', 'HR_ADMIN'),
        (4, '성과 분석', '성과 분석 (/analytics/performance)', 'HR_ADMIN'),
        (5, '출퇴근 분석', '근태 분석 (/analytics/attendance)', 'HR_ADMIN'),
        (6, '이직 분석', '이직률 분석 (/analytics/turnover)', 'HR_ADMIN'),
        (7, '팀 건강도', '팀 건강도 지표 (/analytics/team-health)', 'HR_ADMIN'),
        (8, 'AI 리포트', 'AI 기반 분석 리포트 (Claude) (/analytics/ai-report)', 'HR_ADMIN'),
        (9, '채용 분석', '채용 파이프라인 분석 (/analytics/recruitment)', 'HR_ADMIN'),
        (10, '보상 분석', '보상 분석 (/analytics/compensation)', 'HR_ADMIN'),
        (11, '성별 임금격차', '성별 임금 격차 분석 (/analytics/gender-pay-gap)', 'HR_ADMIN'),
        (12, '이탈 분석', '이탈 분석 (/analytics/attrition)', 'HR_ADMIN'),
        (13, '예측 모델', '예측 분석 (/analytics/predictive)', 'HR_ADMIN'),
        (14, '예측 모델', '직원별 예측 (/analytics/predictive/[employeeId])', 'HR_ADMIN'),
        (15, '리포트', '커스텀 리포트 (/analytics/report)', 'HR_ADMIN'),
        (16, '대시보드', '대시보드 비교 (/dashboard/compare)', 'HR_ADMIN'),
        (17, 'TTM', 'TTM (Trailing Twelve Months) 기본값 작동', 'HR_ADMIN'),
        (18, 'KRW 환산', '해외 법인 통화 KRW 자동 환산', 'HR_ADMIN'),
    ]
    write_checklist_rows(ws, rows)
    add_dropdown(ws, f'E2:E{len(rows) + 1}', ['Y', 'N', 'NA'])
    add_dropdown(ws, f'F2:F{len(rows) + 1}', ['Critical', 'Major', 'Minor', '개선'])
    return len(rows)


def sheet_onboarding(wb):
    """온보딩·오프보딩 — EXPANDED scenario with 4 게이트 + do-not-rehire"""
    ws = wb.create_sheet('온보딩·오프보딩')
    ws.sheet_properties.tabColor = SCENARIO_TAB
    apply_header(ws, SCENARIO_HEADERS, SCENARIO_WIDTHS)

    rows = [
        (1, '온보딩 마일스톤 (DAY_1)', 'HR_ADMIN', '신규 입사자 등록 후',
         '대상: 테스트직원 (인사관리 Step 6)',
         '/onboarding 진입 > 대상 직원 선택 > DAY_1 태스크 확인',
         'DAY_1/7/30/90 4단계 마일스톤이 자동 생성되고 상태가 표시됨'),
        (2, '온보딩 체크인 (직원)', 'EMPLOYEE (신입)', '온보딩 진행 중', '',
         '나의 공간 > 나의 온보딩 > 체크인 (/onboarding/checkin)',
         '신입사원 주간 체크인 폼이 표시되고 제출 가능'),
        (3, '온보딩 체크인 대시보드 (HR)', 'HR_ADMIN', 'Step 2 완료', '',
         '/onboarding/checkins 진입',
         'HR 관리자 체크인 대시보드에 신입 응답 집계 표시'),
        (4, '오프보딩 시작 → 4 게이트 생성', 'HR_ADMIN',
         '퇴사 예정자 존재 (EDGE-007)', '오퇴사 (EDGE-007)',
         '직원 상세 > 오프보딩 시작',
         '4 게이트 체크리스트 자동 생성 (IT계정/퇴직면담/인수인계/자산반납) + D-day 배지 표시'),
        (5, '오프보딩 게이트 1: IT 계정 회수', 'HR_ADMIN', 'Step 4 완료', '',
         '오프보딩 상세 > IT 계정 게이트 진입 > 회수 완료 체크',
         'IT 계정 게이트 완료. 미완료 시 후속 게이트 진행 불가 (강제 집행 엔진)'),
        (6, '오프보딩 게이트 2: 퇴직 면담', 'HR_ADMIN', 'Step 5 완료', '',
         '/offboarding/exit-interviews 진입 > 퇴직 면담 기록 (5-record privacy)',
         '퇴직 면담이 익명·익명 처리되어 기록됨 (5명 이상 모집 시에만 통계 공개)'),
        (7, '오프보딩 게이트 3: 인수인계 워크플로', 'HR_ADMIN', 'Step 6 완료',
         '인수자: 정다은',
         '오프보딩 상세 > 인수인계 게이트 > 인수자 지정 > 인계 태스크 생성',
         '인수자 지정 + 인계 태스크 필터 뷰 + 진행률 바 표시'),
        (8, '오프보딩 게이트 4: 자산 반납', 'HR_ADMIN', 'Step 7 완료', '',
         '오프보딩 상세 > 자산 반납 게이트 > 6-country 규칙 확인',
         '자산 반납 게이트 완료 (6-country 규칙 적용 — KR/CN/US/MX/VN/RU 별)'),
        (9, '4 게이트 완료 → COMPLETED 전이', 'HR_ADMIN', '모든 게이트 완료', '',
         '오프보딩 상세에서 최종 완료 버튼 클릭',
         '오프보딩 COMPLETED 상태로 전환. 게이트 미완료 시 전이 차단 검증'),
        (10, '재고용 방지 (do-not-rehire) 토글', 'HR_ADMIN', 'Step 9 완료',
         '사유: 정책 위반',
         '오프보딩 완료 후 do-not-rehire 토글 ON + 사유 기록',
         'do-not-rehire 플래그가 활성화되고 채용 모듈에서 차단 가능'),
        (11, '채용 시 재고용 차단 검증', 'HR_ADMIN', 'Step 10 완료',
         '동일 인물 채용 시도',
         '채용 > 새 지원자 등록 시 do-not-rehire 직원 이메일/이름 입력',
         '중복 체크 시 경고 또는 차단 표시'),
        (12, 'OffboardingDocument 수집', 'HR_ADMIN', '오프보딩 진행 중', '',
         '오프보딩 상세 > 문서 수집 CRUD',
         'CONSENT/HANDOVER/EXIT/NDA 4 종 문서 업로드 + 관리 가능'),
        (13, '퇴직 분석 대시보드', 'HR_ADMIN', '퇴직 데이터 누적', '',
         '/analytics/turnover 또는 오프보딩 분석',
         '이직 원인/만족도/재직기간/월별 트렌드가 시각화됨'),
        (14, '내 온보딩 (직원 뷰)', 'EMPLOYEE (신입)', '온보딩 진행 중', '',
         '/onboarding/me 진입',
         '직원 본인의 온보딩 진행 상황 + 마일스톤 + 태스크 표시'),
        (15, '내 오프보딩 (직원 뷰)', 'EMPLOYEE (퇴사 예정)', '오프보딩 시작됨', '',
         '/my/offboarding 진입',
         '직원 본인의 오프보딩 진행 상황 + 4 게이트 진행률 표시'),
    ]
    write_scenario_rows(ws, rows, mark_new={5, 6, 7, 8, 9, 10, 11, 12, 13})
    add_dropdown(ws, f'I2:I{len(rows) + 1}', ['Pass', 'Fail', 'Block', 'N/A'])
    add_conditional_pass_fail(ws, f'I2:I{len(rows) + 1}')
    return len(rows)


def sheet_compliance(wb):
    """규정준수 — checklist"""
    ws = wb.create_sheet('규정준수')
    ws.sheet_properties.tabColor = CHECKLIST_TAB
    apply_header(ws, CHECKLIST_HEADERS, CHECKLIST_WIDTHS)

    rows = [
        (1, '컴플라이언스', '컴플라이언스 허브 (/compliance)', 'HR_ADMIN'),
        (2, 'GDPR', 'GDPR 관리 (/compliance/gdpr)', 'HR_ADMIN'),
        (3, 'DPIA', 'DPIA 평가 (/compliance/dpia)', 'HR_ADMIN'),
        (4, '데이터 보존', '데이터 보존 정책 (/compliance/data-retention)', 'HR_ADMIN'),
        (5, 'PII 감사', 'PII 감사 로그 (/compliance/pii-audit)', 'HR_ADMIN'),
        (6, '한국 규정', '한국 노동법 준수 (/compliance/kr)', 'HR_ADMIN'),
        (7, '중국 규정', '중국 노동법 준수 (/compliance/cn)', 'HR_ADMIN (hr@ctr-cn.com)'),
        (8, '러시아 규정', '러시아 노동법 준수 (/compliance/ru)', 'HR_ADMIN'),
        (9, 'RLS', 'Row-Level Security (RLS P1) — 회사 격리', 'HR_ADMIN (cross-company)'),
        (10, 'PII 로깅', 'PII access 자동 로깅', 'HR_ADMIN'),
    ]
    write_checklist_rows(ws, rows)
    add_dropdown(ws, f'E2:E{len(rows) + 1}', ['Y', 'N', 'NA'])
    add_dropdown(ws, f'F2:F{len(rows) + 1}', ['Critical', 'Major', 'Minor', '개선'])
    return len(rows)


def sheet_my_space(wb):
    """My Space — checklist"""
    ws = wb.create_sheet('My Space')
    ws.sheet_properties.tabColor = CHECKLIST_TAB
    apply_header(ws, CHECKLIST_HEADERS, CHECKLIST_WIDTHS)

    rows = [
        (1, 'My Space', '내 할일 (/my/tasks)', 'EMPLOYEE'),
        (2, 'My Space', '내 출퇴근 (/attendance — EMPLOYEE 뷰)', 'EMPLOYEE'),
        (3, 'My Space', '내 휴가 (/my/leave)', 'EMPLOYEE'),
        (4, 'My Space', '내 휴직 (/leave-of-absence — EMPLOYEE 뷰)', 'EMPLOYEE'),
        (5, 'My Space', '내 급여명세서 (/payroll/me)', 'EMPLOYEE'),
        (6, 'My Space', '내 급여명세서 상세 (/payroll/me/[runId])', 'EMPLOYEE'),
        (7, 'My Space', '내 복리후생 (/my/benefits)', 'EMPLOYEE'),
        (8, 'My Space', '내 연말정산 (/my/year-end)', 'EMPLOYEE'),
        (9, 'My Space', '내 목표 (/performance/my-goals)', 'EMPLOYEE'),
        (10, 'My Space', '내 분기리뷰 (/performance/my-quarterly-review)', 'EMPLOYEE'),
        (11, 'My Space', '내 스킬 (/my/skills)', 'EMPLOYEE'),
        (12, 'My Space', '내 교육 (/my/training)', 'EMPLOYEE'),
        (13, 'My Space', '내 인정/칭찬 (/performance/recognition)', 'EMPLOYEE'),
        (14, 'My Space', '내 문서 (/my/documents)', 'EMPLOYEE'),
        (15, 'My Space', '내 프로필 (/my/profile)', 'EMPLOYEE'),
        (16, 'My Space', '내 알림설정 (/my/settings/notifications)', 'EMPLOYEE'),
        (17, 'My Space', '내 총보상 (/my/total-rewards) — Total Rewards Statement', 'EMPLOYEE'),
        (18, 'My Space', '내 온보딩 (/onboarding/me)', 'EMPLOYEE (신입)'),
        (19, 'My Space', '내 오프보딩 (/my/offboarding)', 'EMPLOYEE (퇴사 예정)'),
        (20, 'My Space', '내 내부공모 (/my/internal-jobs)', 'EMPLOYEE'),
    ]
    write_checklist_rows(ws, rows)
    add_dropdown(ws, f'E2:E{len(rows) + 1}', ['Y', 'N', 'NA'])
    add_dropdown(ws, f'F2:F{len(rows) + 1}', ['Critical', 'Major', 'Minor', '개선'])
    return len(rows)


def sheet_v2_home(wb):
    """V2 홈 — NEW SHEET — scenario for role-based homepages"""
    ws = wb.create_sheet('V2 홈')
    ws.sheet_properties.tabColor = SCENARIO_TAB
    apply_header(ws, SCENARIO_HEADERS, SCENARIO_WIDTHS)

    rows = [
        (1, 'SUPER_ADMIN 홈', 'SUPER_ADMIN (super@ctr.co.kr)', '로그인 완료', '',
         '로그인 후 자동으로 /home으로 리다이렉트됨',
         'SUPER_ADMIN 전용 홈 (13개 법인 cross-company 위젯 + 전사 KPI) 표시'),
        (2, 'HR_ADMIN 홈 (HrAdminHomeV2)', 'HR_ADMIN (hr@ctr.co.kr)', '로그인 완료', '',
         '/home 진입',
         'HrAdminHomeV2 카드 — 결재 inbox 요약, 온보딩/오프보딩 tracker, 휴가 통계, recognition feed'),
        (3, 'MANAGER 홈 (ManagerHomeV2)', 'MANAGER (manager@ctr.co.kr)', '로그인 완료', '',
         '/home 진입',
         'ManagerHomeV2 — 팀원 현황, 결재 대기, 1:1 일정, 팀 성과 요약'),
        (4, 'EXECUTIVE 홈 (ExecutiveHomeV2)', 'EXECUTIVE / SUPER_ADMIN', '로그인 완료', '',
         '/home 진입',
         '경영진 요약 (HR KPI + 채용 퍼널 + 인력 트렌드)'),
        (5, 'EMPLOYEE 홈 (EmployeeHomeV2)', 'EMPLOYEE (employee-a@ctr.co.kr)', '로그인 완료', '',
         '/home 진입',
         'EmployeeHomeV2 — 내 할일, 휴가 잔액, 목표 진행률, 인정 피드, 출퇴근 위젯'),
        (6, '시간대별 인사말 (hydration-safe)', 'EMPLOYEE', '로그인 완료', '',
         '홈에 진입 시 시간대 인사말 확인 (아침/낮/저녁)',
         '현지 시간대 기반 인사말이 hydration 오류 없이 표시됨 (Session 181 fix)'),
        (7, '온보딩/오프보딩 tracker row 클릭', 'HR_ADMIN', 'tracker row 표시', '',
         '홈의 tracker row 클릭',
         '해당 onboarding/offboarding 상세 페이지로 정확히 이동 (Session 182 fix)'),
        (8, '알림 위젯', 'all roles', '신규 알림 존재', '',
         '홈의 알림 위젯 또는 /notifications 클릭',
         '신규 알림 N건 카운트 + 알림 목록 표시'),
        (9, 'SW 자동 갱신 toast', 'all roles', '신규 배포 직후', '',
         '신규 배포 후 페이지 진입',
         '"새 버전이 있습니다 / 새로고침" toast 표시 + 1-click 새로고침 정상 동작 (Session 199 prod 검증)'),
    ]
    write_scenario_rows(ws, rows, mark_new=set(range(1, 10)))  # all new
    add_dropdown(ws, f'I2:I{len(rows) + 1}', ['Pass', 'Fail', 'Block', 'N/A'])
    add_conditional_pass_fail(ws, f'I2:I{len(rows) + 1}')
    return len(rows)


def sheet_notifications(wb):
    """알림 — NEW SHEET — checklist"""
    ws = wb.create_sheet('알림')
    ws.sheet_properties.tabColor = CHECKLIST_TAB
    apply_header(ws, CHECKLIST_HEADERS, CHECKLIST_WIDTHS)

    rows = [
        (1, '알림 센터', '알림 목록 (/notifications)', 'all roles'),
        (2, '알림 센터', '읽음/안읽음 토글', 'all roles'),
        (3, '알림 센터', '카테고리 필터 (휴가/결재/평가 등)', 'all roles'),
        (4, '알림 발송', '결재 대기 알림 (Teams + 이메일 + 푸시)', 'MANAGER'),
        (5, '알림 발송', '휴가 시작 D-7/D-3/D-1 알림', 'EMPLOYEE'),
        (6, '알림 발송', '복귀 알림 (LOA D-7/D-3/D-1)', 'EMPLOYEE / HR_ADMIN'),
        (7, '알림 발송', '평가 마감 알림 (eval-reminder cron)', 'EMPLOYEE / MANAGER'),
        (8, '알림 발송', 'nudge cron (매니저급)', 'MANAGER+'),
        (9, '알림 발송', 'HR 챗봇 에스컬레이션', 'HR_ADMIN'),
        (10, '알림 i18n', '5 locale (ko/en/zh/vi/es) 알림 본문', 'multi-locale'),
        (11, '알림 설정', '내 알림 설정 (/my/settings/notifications)', 'EMPLOYEE'),
    ]
    write_checklist_rows(ws, rows)
    add_dropdown(ws, f'E2:E{len(rows) + 1}', ['Y', 'N', 'NA'])
    add_dropdown(ws, f'F2:F{len(rows) + 1}', ['Critical', 'Major', 'Minor', '개선'])
    return len(rows)


def sheet_edge_cases(wb):
    """엣지케이스 — 30 EDGE personas"""
    ws = wb.create_sheet('엣지케이스')
    ws.sheet_properties.tabColor = EDGE_TAB
    apply_header(ws, SCENARIO_HEADERS, SCENARIO_WIDTHS)

    rows = [
        (1, '수습 진행 중', 'HR_ADMIN', 'EDGE-001 시드', '김수습 (EDGE-001)',
         '직원 목록 검색 > 상세 페이지', '수습 배지 + 만료일(2026-06-01) 표시'),
        (2, '수습 만료 경고', 'HR_ADMIN', 'EDGE-002 시드', '박만료 (EDGE-002)',
         '직원 목록 검색', '수습 기간 만료 경고 (2026-01-01 지남) 표시'),
        (3, '계약 만료 임박', 'HR_ADMIN', 'EDGE-003 시드', '이계약 (EDGE-003)',
         '직원 목록 또는 대시보드', 'D-30 알림 + 계약 갱신/종료 조치 안내'),
        (4, '계약 만료 직원', 'HR_ADMIN', 'EDGE-004 시드', '최만기 (EDGE-004)',
         '직원 목록 검색', '계약 만료 상태 (만료일: 2026-03-31) 표시'),
        (5, '육아휴직 중', 'HR_ADMIN', 'EDGE-005 시드', '정육아 (EDGE-005)',
         '직원 목록 + 근태/급여 관리', '출퇴근/급여 대상에서 제외 + 휴직 배지'),
        (6, '복귀 예정 휴직자', 'HR_ADMIN', 'EDGE-006 시드', '한복귀 (EDGE-006)',
         '직원 목록 또는 휴가/휴직 관리', '복귀 예정 알림 + 복귀 처리 버튼'),
        (7, '퇴사 예정자 오프보딩', 'HR_ADMIN', 'EDGE-007 시드', '오퇴사 (EDGE-007)',
         '직원 목록 + 오프보딩', '4 게이트 체크리스트 생성됨 (퇴사일 2026-05-06)'),
        (8, '퇴직 완료 접근 차단', 'HR_ADMIN', 'EDGE-008 시드', '유퇴직 (EDGE-008)',
         '활성 직원 목록 필터', '활성 목록에서 제외 + 로그인 차단'),
        (9, '겸직자 (2개 법인)', 'HR_ADMIN', 'EDGE-009 시드', '강겸직 (EDGE-009)',
         '직원 상세 > 발령 정보', '2개 법인 발령 모두 표시'),
        (10, '다직위자 (3개 직위)', 'HR_ADMIN', 'EDGE-010 시드', '임다직 (EDGE-010)',
         '직원 상세', '3개 직위 모두 표시 + 주직위 구분'),
        (11, '매니저 없는 직원 (CEO)', 'HR_ADMIN', 'EDGE-011 시드', '윤대표 (EDGE-011)',
         '직원 상세 + 결재 라인', '상위 보고자 없이 정상 표시. 결재 스킵 또는 HR 직접'),
        (12, 'EMPLOYEE 타인 급여 차단', 'EMPLOYEE (employee-a)', '로그인 완료',
         'URL: /payroll/me/[다른직원RunID]', '주소창에 직접 입력', '403 또는 본인 급여만 조회 가능'),
        (13, '타 법인 데이터 격리', 'HR_ADMIN (CTR)', '로그인 완료', '',
         'CTR HR_ADMIN으로 CTR-CN 직원 접근 시도', 'CTR-CN 데이터 미표시 (법인 격리)'),
        (14, '부서 미배정 신규 입사자', 'HR_ADMIN', 'EDGE-012 시드', '신입사 (EDGE-012)',
         '직원 목록', '부서 미배정 상태로 오류 없이 표시 + 부서 배정 필요 알림'),
        (15, '법인 간 전적', 'HR_ADMIN', 'EDGE-013 시드', '배전적 (EDGE-013)',
         '직원 상세', '법인 전적(CTR→CTR-CN) 이력 발령 이력에 기록'),
        (16, 'Grade 없음', 'HR_ADMIN', 'EDGE-014 시드', 'Alex NoGrade (EDGE-014)',
         '직원 상세', 'Grade 미할당 상태 오류 없이 표시'),
        (17, '연봉 밴드 초과자', 'HR_ADMIN', 'EDGE-015 시드', '고초과 (EDGE-015)',
         '보상 관리', '밴드 상한 초과 경고'),
        (18, '연봉 밴드 미달자', 'HR_ADMIN', 'EDGE-016 시드', '하미달 (EDGE-016)',
         '보상 관리', '밴드 하한 미달 경고'),
        (19, '무급 인턴', 'HR_ADMIN', 'EDGE-017 시드', '무급실 (EDGE-017)',
         '급여 실행', '급여 대상 제외 또는 0원'),
        (20, '빈번한 급여 변경', 'HR_ADMIN', 'EDGE-018 시드', '변봉급 (EDGE-018)',
         '급여 이상 검토', '잦은 변경 이력 이상 항목으로 감지'),
        (21, '비정기 조정 대기', 'HR_ADMIN', 'EDGE-019 시드', '승대기 (EDGE-019)',
         '비정기 조정 목록', '승인 대기 건 표시'),
        (22, '재수정 요청 직원', 'HR_ADMIN', 'EDGE-020 시드', '재수정 (EDGE-020)',
         '결재함', '재수정 요청 흐름 처리'),
        (23, '연차 잔액 0일', 'HR_ADMIN', 'EDGE-021 시드', '공소진 (EDGE-021)',
         '휴가 관리 + 추가 신청 시도', '잔액 0일 + 차단 또는 마이너스 경고'),
        (24, '연차 마이너스 잔액', 'HR_ADMIN', 'EDGE-022 시드', '마이너 (EDGE-022)',
         '휴가 관리', '마이너스 잔액 표시 + 급여 공제 연동 확인'),
        (25, '빈번한 지각', 'HR_ADMIN', 'EDGE-023 시드', '지각왕 (EDGE-023)',
         '근태 관리 이력', '빈번한 지각 패턴 이상 항목으로 감지'),
        (26, '52시간 초과 위험', 'HR_ADMIN', 'EDGE-024 시드', '과로자 (EDGE-024)',
         '근태 관리 또는 대시보드', '주 52시간 초과 위험 경고'),
        (27, '복합 휴가 직원', 'HR_ADMIN', 'EDGE-025 시드', '복합휴 (EDGE-025)',
         '휴가 관리', '복수 휴가 유형 동시 적용 정보 정확히 표시'),
        (28, '목표 수정 직원', 'HR_ADMIN', 'EDGE-026 시드', '목수정 (EDGE-026)',
         '성과 관리', 'GoalRevision 이력 표시'),
        (29, '분기 리뷰 직원', 'HR_ADMIN', 'EDGE-027 시드', '분리뷰 (EDGE-027)',
         '분기 리뷰 목록', 'QuarterlyReview 듀얼 제출 이력 확인'),
        (30, '등급 하향 직원', 'HR_ADMIN', 'EDGE-028 시드', '등하향 (EDGE-028)',
         '캘리브레이션', 'DnD 등급 하향 조정 결과 보존'),
        (31, '잘못된 URL 직접 입력', 'EMPLOYEE', '로그인 완료', 'URL: /존재하지않는페이지',
         '주소창에 존재하지 않는 경로 입력', '404 페이지 표시 또는 홈 리다이렉트'),
        (32, '동시 세션', 'EMPLOYEE', '로그인 완료', '2개 브라우저 동시 로그인',
         '두 브라우저에서 동일 계정 동시 작업', '두 세션 정상 동작 또는 보안 정책에 따라 이전 세션 만료'),
    ]
    write_scenario_rows(ws, rows, mark_new={28, 29, 30})
    add_dropdown(ws, f'I2:I{len(rows) + 1}', ['Pass', 'Fail', 'Block', 'N/A'])
    add_conditional_pass_fail(ws, f'I2:I{len(rows) + 1}')
    return len(rows)


def sheet_summary(wb, counts):
    """집계 — auto COUNTIF formulas"""
    ws = wb.create_sheet('집계')
    ws.sheet_properties.tabColor = META_TAB

    ws['A1'] = 'CTR HR Hub UAT v2 테스트 집계'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:I1')

    # 모듈별 시나리오 집계
    ws['A3'] = '시나리오·체크리스트 모듈별 집계'
    ws['A3'].font = SECTION_FONT

    headers = ['모듈', '총 항목수', 'Pass', 'Fail', 'Block', 'N/A', '미입력', '통과율(%)']
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = ALIGN_CENTER
        c.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # Module rows — link to source sheets with COUNTIF
    modules = [
        ('인사관리', 'I', 2, counts['인사관리'] + 1),
        ('출퇴근', 'I', 2, counts['출퇴근'] + 1),
        ('휴가관리', 'I', 2, counts['휴가관리'] + 1),
        ('휴직', 'I', 2, counts['휴직'] + 1),
        ('급여', 'I', 2, counts['급여'] + 1),
        ('결재·승인', 'I', 2, counts['결재·승인'] + 1),
        ('마스터데이터·설정', 'I', 2, counts['마스터데이터·설정'] + 1),
        ('성과', 'I', 2, counts['성과'] + 1),
        ('채용', 'E', 2, counts['채용'] + 1),  # checklist uses E (확인)
        ('보상', 'I', 2, counts['보상'] + 1),
        ('인사이트·분석', 'E', 2, counts['인사이트·분석'] + 1),
        ('온보딩·오프보딩', 'I', 2, counts['온보딩·오프보딩'] + 1),
        ('규정준수', 'E', 2, counts['규정준수'] + 1),
        ('My Space', 'E', 2, counts['My Space'] + 1),
        ('V2 홈', 'I', 2, counts['V2 홈'] + 1),
        ('알림', 'E', 2, counts['알림'] + 1),
        ('엣지케이스', 'I', 2, counts['엣지케이스'] + 1),
    ]

    for i, (mod, col, start, end) in enumerate(modules):
        r = 5 + i
        ws.cell(row=r, column=1, value=mod).font = BODY_FONT
        ws.cell(row=r, column=2, value=end - start + 1).font = BODY_FONT
        if col == 'I':
            range_str = f"'{mod}'!{col}{start}:{col}{end}"
            ws.cell(row=r, column=3, value=f'=COUNTIF({range_str},"Pass")')
            ws.cell(row=r, column=4, value=f'=COUNTIF({range_str},"Fail")')
            ws.cell(row=r, column=5, value=f'=COUNTIF({range_str},"Block")')
            ws.cell(row=r, column=6, value=f'=COUNTIF({range_str},"N/A")')
            ws.cell(row=r, column=7, value=f'=B{r}-C{r}-D{r}-E{r}-F{r}')
            ws.cell(row=r, column=8, value=f'=IFERROR(C{r}/B{r}*100,0)').number_format = '0.0'
        else:  # checklist: Y/N/NA
            range_str = f"'{mod}'!{col}{start}:{col}{end}"
            ws.cell(row=r, column=3, value=f'=COUNTIF({range_str},"Y")')  # Pass-equivalent
            ws.cell(row=r, column=4, value=f'=COUNTIF({range_str},"N")')  # Fail-equivalent
            ws.cell(row=r, column=5, value=0)
            ws.cell(row=r, column=6, value=f'=COUNTIF({range_str},"NA")')
            ws.cell(row=r, column=7, value=f'=B{r}-C{r}-D{r}-E{r}-F{r}')
            ws.cell(row=r, column=8, value=f'=IFERROR(C{r}/B{r}*100,0)').number_format = '0.0'
        for col_idx in range(1, 9):
            ws.cell(row=r, column=col_idx).border = BORDER
            ws.cell(row=r, column=col_idx).font = BODY_FONT

    # 합계
    total_row = 5 + len(modules)
    ws.cell(row=total_row, column=1, value='합계').font = HEADER_FONT
    ws.cell(row=total_row, column=1).fill = HEADER_FILL
    for col_idx in range(2, 8):
        col_letter = get_column_letter(col_idx)
        ws.cell(row=total_row, column=col_idx,
                value=f'=SUM({col_letter}5:{col_letter}{total_row - 1})')
        ws.cell(row=total_row, column=col_idx).font = HEADER_FONT
        ws.cell(row=total_row, column=col_idx).fill = HEADER_FILL
    ws.cell(row=total_row, column=8, value=f'=IFERROR(C{total_row}/B{total_row}*100,0)').number_format = '0.0'
    ws.cell(row=total_row, column=8).font = HEADER_FONT
    ws.cell(row=total_row, column=8).fill = HEADER_FILL
    for col_idx in range(1, 9):
        ws.cell(row=total_row, column=col_idx).border = BORDER


def attach_part2(wb, counts):
    """Append remaining sheets and update counts dict."""
    counts['성과'] = sheet_performance(wb)
    counts['채용'] = sheet_recruitment(wb)
    counts['보상'] = sheet_compensation(wb)
    counts['인사이트·분석'] = sheet_insights(wb)
    counts['온보딩·오프보딩'] = sheet_onboarding(wb)
    counts['규정준수'] = sheet_compliance(wb)
    counts['My Space'] = sheet_my_space(wb)
    counts['V2 홈'] = sheet_v2_home(wb)
    counts['알림'] = sheet_notifications(wb)
    counts['엣지케이스'] = sheet_edge_cases(wb)
    sheet_summary(wb, counts)
