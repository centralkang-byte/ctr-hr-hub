#!/usr/bin/env python3
"""
CTR HR Hub UAT 피드백 결과 v2 generator
Run: python3 scripts/uat/build-feedback-v2.py
Output: docs/uat/UAT_피드백_결과_v2.xlsx

v2 changes vs v1:
- COUNTIFS auto-formulas applied to 모듈별 집계 / 심각도별 집계 / 일별 트리아지
- 휴직 모듈 추가 (was missing in v1)
- Module labels aligned with 12-manual SSOT
- Data validation dropdowns for 모듈/심각도/상태 columns
- Pass/Fail/Block 조건부 서식
- UAT 일자 baseline = 2026-05-19 (Day 1) — UAT 가이드 §2 기준
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from pathlib import Path
from datetime import date, timedelta

REPO_ROOT = Path(__file__).resolve().parents[2]
OUT_PATH = REPO_ROOT / 'docs/uat/UAT_피드백_결과_v2.xlsx'

# ============================================================
# Styles
# ============================================================
FONT_NAME = 'Malgun Gothic'
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
HEADER_FILL = PatternFill('solid', fgColor='2E75B6')
HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color='FFFFFF')
SECTION_FILL = PatternFill('solid', fgColor='DDEBF7')
SECTION_FONT = Font(name=FONT_NAME, size=11, bold=True, color='1F4E78')
TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True, color='1F4E78')
BODY_FONT = Font(name=FONT_NAME, size=11)
ALIGN_TOP_WRAP = Alignment(vertical='top', wrap_text=True)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

META_TAB = 'A6A6A6'
LOG_TAB = 'C00000'
TRIAGE_TAB = 'ED7D31'
SUMMARY_TAB = '70AD47'
JUDGMENT_TAB = '7030A0'

# ============================================================
# Module list (aligned with 12-manual SSOT + 휴직 추가)
# ============================================================
MODULES = [
    '인사관리', '출퇴근', '휴가관리', '휴직', '급여', '결재·승인',
    '마스터데이터·설정', '성과', '채용', '보상', '인사이트·분석',
    '온보딩·오프보딩', '규정준수', 'My Space', 'V2 홈', '알림', '엣지케이스', '기타',
]

SEVERITIES = ['Critical', 'Major', 'Minor', '개선']
STATUSES = ['신규', '확인중', '수정중', '수정완료', '재테스트', '종료', '보류']
GO_LIVE_OPTS = ['차단', '비차단', '미정']
ACCEPT_OPTS = ['수용', '보류', '미정']

# UAT Day 1 baseline (per UAT 가이드 v2 §2 — 2주 일정 가정)
UAT_DAY1 = date(2026, 5, 19)


# ============================================================
# Helpers
# ============================================================
def apply_header(ws, headers, widths, row=1):
    for col_idx, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = ALIGN_CENTER
        c.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = f'A{row + 1}'


# ============================================================
# Sheet 1: 리포트 작성법
# ============================================================
def sheet_guide(wb):
    ws = wb.create_sheet('리포트 작성법')
    ws.sheet_properties.tabColor = META_TAB

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 80

    ws['A1'] = 'CTR HR Hub UAT v2 피드백 리포트 작성 가이드'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:B1')

    ws['A3'] = '심각도 기준'
    ws['A3'].font = SECTION_FONT

    rows_severity = [
        ('심각도', '정의 / 예시'),
        ('Critical', '시스템 사용 불가, 데이터 손실/유출 (예: 로그인 불가, 급여 오계산, 타 법인 데이터 노출)'),
        ('Major', '핵심 업무 수행 방해 (예: 휴가 신청 실패, 결재 안됨, 직원 등록 오류)'),
        ('Minor', '불편하지만 우회 가능 (예: UI 깨짐, 번역 누락, 느린 로딩)'),
        ('개선', '기능 개선 제안 (예: UX 개선, 필터 추가, 라벨 변경)'),
    ]
    for i, (a, b) in enumerate(rows_severity):
        r = 4 + i
        ws.cell(row=r, column=1, value=a)
        ws.cell(row=r, column=2, value=b)
        if i == 0:
            ws.cell(row=r, column=1).font = HEADER_FONT
            ws.cell(row=r, column=1).fill = HEADER_FILL
            ws.cell(row=r, column=2).font = HEADER_FONT
            ws.cell(row=r, column=2).fill = HEADER_FILL
        else:
            ws.cell(row=r, column=1).font = BODY_FONT
            ws.cell(row=r, column=2).font = BODY_FONT
        for col in (1, 2):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).alignment = ALIGN_TOP_WRAP

    ws['A11'] = '작성 예시'
    ws['A11'].font = SECTION_FONT

    example = [
        ('필드', '예시 값'),
        ('이슈ID', 'FB-007'),
        ('모듈', '휴가관리'),
        ('시나리오/항목#', '휴가관리 Step 7 (반차 신청)'),
        ('설명', '반차 신청 시 잔여일수가 1일 차감됨 (0.5일이어야 함)'),
        ('기대결과', '반차 신청 후 잔여일수 14.5일'),
        ('실제결과', '잔여일수 14일로 표시'),
        ('심각도', 'Major'),
        ('스크린샷', '(캡처 이미지 첨부)'),
        ('테스터', '한지영'),
        ('발견일', '2026-05-20'),
        ('상태', '신규'),
        ('Go-live 차단', '비차단'),
    ]
    for i, (a, b) in enumerate(example):
        r = 12 + i
        ws.cell(row=r, column=1, value=a)
        ws.cell(row=r, column=2, value=b)
        if i == 0:
            ws.cell(row=r, column=1).font = HEADER_FONT
            ws.cell(row=r, column=1).fill = HEADER_FILL
            ws.cell(row=r, column=2).font = HEADER_FONT
            ws.cell(row=r, column=2).fill = HEADER_FILL
        else:
            ws.cell(row=r, column=1).font = BODY_FONT
            ws.cell(row=r, column=2).font = BODY_FONT
        for col in (1, 2):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).alignment = ALIGN_TOP_WRAP

    base_row = 12 + len(example) + 2
    ws.cell(row=base_row, column=1, value='리포트 규칙').font = SECTION_FONT
    rules = [
        '1. 스크린샷은 필수 (Print Screen → Excel에 붙여넣기)',
        '2. URL은 브라우저 주소창에서 복사하여 비고에 기록',
        '3. 재현 스텝은 최대한 구체적으로 기록 (수행한 클릭 순서)',
        '4. 이미 보고된 건은 중복 기입하지 않고, 기존 이슈ID를 비고에 참조',
        '5. Critical 발견 시 즉시 UAT 관리자에게 구두 보고',
        '6. v1에 없던 신규 기능 (V2 홈, Compensation Letter, Off-Cycle Comp, GoalRevision, QuarterlyReview 듀얼제출, 캘리브레이션 DnD, 인수인계 4 게이트, do-not-rehire, dept_head 결재, 알림 센터, 지정연차) 검증 시 매뉴얼 §1~§10을 먼저 읽고 기대 결과를 명확히 한 후 보고',
    ]
    for i, r_text in enumerate(rules):
        r = base_row + 1 + i
        c = ws.cell(row=r, column=1, value=r_text)
        c.font = BODY_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)


# ============================================================
# Sheet 2: 피드백 로그
# ============================================================
def sheet_log(wb):
    ws = wb.create_sheet('피드백 로그')
    ws.sheet_properties.tabColor = LOG_TAB

    headers = [
        '이슈ID', '모듈', '시나리오/항목#', '설명', '기대결과', '실제결과', '심각도',
        '스크린샷', '테스터', '발견일', '상태', '재테스트 상태', '결함 경과일',
        '담당자', '수용/보류', 'Go-live 차단', '해결일', '비고',
    ]
    widths = [10, 18, 22, 35, 30, 30, 12, 14, 12, 12, 12, 14, 12, 14, 12, 14, 12, 24]
    apply_header(ws, headers, widths)

    # 100 empty rows with auto ID
    for i in range(1, 101):
        r = i + 1
        ws.cell(row=r, column=1, value=f'FB-{i:03d}').font = BODY_FONT
        # 결함 경과일 = TODAY() - 발견일 (if 발견일 있을 때)
        ws.cell(row=r, column=13, value=f'=IFERROR(IF(J{r}="","",TODAY()-J{r}),"")').font = BODY_FONT
        for col_idx in range(1, 19):
            ws.cell(row=r, column=col_idx).border = BORDER
            ws.cell(row=r, column=col_idx).alignment = ALIGN_TOP_WRAP
        ws.row_dimensions[r].height = 30

    # Data validations
    dv_module = DataValidation(type='list', formula1=f'"{",".join(MODULES)}"', allow_blank=True)
    dv_module.add('B2:B101')
    ws.add_data_validation(dv_module)

    dv_severity = DataValidation(type='list', formula1=f'"{",".join(SEVERITIES)}"', allow_blank=True)
    dv_severity.add('G2:G101')
    ws.add_data_validation(dv_severity)

    dv_status = DataValidation(type='list', formula1=f'"{",".join(STATUSES)}"', allow_blank=True)
    dv_status.add('K2:K101')
    ws.add_data_validation(dv_status)

    dv_retest = DataValidation(type='list', formula1='"Pass,Fail,재테스트 대기"', allow_blank=True)
    dv_retest.add('L2:L101')
    ws.add_data_validation(dv_retest)

    dv_accept = DataValidation(type='list', formula1=f'"{",".join(ACCEPT_OPTS)}"', allow_blank=True)
    dv_accept.add('O2:O101')
    ws.add_data_validation(dv_accept)

    dv_block = DataValidation(type='list', formula1=f'"{",".join(GO_LIVE_OPTS)}"', allow_blank=True)
    dv_block.add('P2:P101')
    ws.add_data_validation(dv_block)

    # Conditional formatting for severity column
    ws.conditional_formatting.add(
        'G2:G101',
        CellIsRule(operator='equal', formula=['"Critical"'],
                   fill=PatternFill('solid', fgColor='C00000'),
                   font=Font(name=FONT_NAME, size=11, bold=True, color='FFFFFF')))
    ws.conditional_formatting.add(
        'G2:G101',
        CellIsRule(operator='equal', formula=['"Major"'],
                   fill=PatternFill('solid', fgColor='FFC7CE'),
                   font=Font(name=FONT_NAME, size=11, color='9C0006')))
    ws.conditional_formatting.add(
        'G2:G101',
        CellIsRule(operator='equal', formula=['"Minor"'],
                   fill=PatternFill('solid', fgColor='FFEB9C'),
                   font=Font(name=FONT_NAME, size=11, color='9C5700')))
    ws.conditional_formatting.add(
        'G2:G101',
        CellIsRule(operator='equal', formula=['"개선"'],
                   fill=PatternFill('solid', fgColor='C6EFCE'),
                   font=Font(name=FONT_NAME, size=11, color='006100')))

    # Conditional formatting for status
    ws.conditional_formatting.add(
        'K2:K101',
        CellIsRule(operator='equal', formula=['"종료"'],
                   fill=PatternFill('solid', fgColor='C6EFCE'),
                   font=Font(name=FONT_NAME, size=11, color='006100')))
    ws.conditional_formatting.add(
        'K2:K101',
        CellIsRule(operator='equal', formula=['"신규"'],
                   fill=PatternFill('solid', fgColor='FFC7CE'),
                   font=Font(name=FONT_NAME, size=11, color='9C0006')))


# ============================================================
# Sheet 3: 일별 트리아지 (auto COUNTIFS)
# ============================================================
def sheet_triage(wb):
    ws = wb.create_sheet('일별 트리아지')
    ws.sheet_properties.tabColor = TRIAGE_TAB

    headers = ['Day', '날짜', '신규 발견', '미해결 누적', '수정완료', '재테스트', '보류', 'Critical 미해결', 'Major 미해결']
    widths = [8, 14, 12, 14, 12, 12, 10, 16, 14]
    apply_header(ws, headers, widths)

    for i in range(1, 11):  # Day 1-10
        r = i + 1
        d = UAT_DAY1 + timedelta(days=i - 1)
        date_str = d.strftime('%Y-%m-%d')
        ws.cell(row=r, column=1, value=f'Day {i}').font = BODY_FONT
        ws.cell(row=r, column=2, value=date_str).font = BODY_FONT
        # 신규 발견 = COUNTIFS where 발견일 = this day
        ws.cell(row=r, column=3,
                value=f'=COUNTIFS(\'피드백 로그\'!J:J,"="&DATE({d.year},{d.month},{d.day}))').font = BODY_FONT
        # 미해결 누적 = COUNTIFS 발견일<=this day AND 상태<>"종료"
        ws.cell(row=r, column=4,
                value=f'=COUNTIFS(\'피드백 로그\'!J:J,"<="&DATE({d.year},{d.month},{d.day}),\'피드백 로그\'!K:K,"<>종료")').font = BODY_FONT
        # 수정완료 = 해결일 = this day
        ws.cell(row=r, column=5,
                value=f'=COUNTIFS(\'피드백 로그\'!Q:Q,"="&DATE({d.year},{d.month},{d.day}))').font = BODY_FONT
        # 재테스트 = 재테스트 상태 = "재테스트 대기" AND 발견일 <= this day
        ws.cell(row=r, column=6,
                value=f'=COUNTIFS(\'피드백 로그\'!L:L,"재테스트 대기",\'피드백 로그\'!J:J,"<="&DATE({d.year},{d.month},{d.day}))').font = BODY_FONT
        # 보류 = 상태 = "보류" AND 발견일 <= this day
        ws.cell(row=r, column=7,
                value=f'=COUNTIFS(\'피드백 로그\'!K:K,"보류",\'피드백 로그\'!J:J,"<="&DATE({d.year},{d.month},{d.day}))').font = BODY_FONT
        # Critical 미해결 = 심각도=Critical AND 상태<>종료 AND 발견일<=this day
        ws.cell(row=r, column=8,
                value=f'=COUNTIFS(\'피드백 로그\'!G:G,"Critical",\'피드백 로그\'!K:K,"<>종료",\'피드백 로그\'!J:J,"<="&DATE({d.year},{d.month},{d.day}))').font = BODY_FONT
        # Major 미해결 (핵심 모듈)
        ws.cell(row=r, column=9,
                value=f'=COUNTIFS(\'피드백 로그\'!G:G,"Major",\'피드백 로그\'!K:K,"<>종료",\'피드백 로그\'!J:J,"<="&DATE({d.year},{d.month},{d.day}))').font = BODY_FONT
        for col in range(1, 10):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).alignment = ALIGN_TOP_WRAP


# ============================================================
# Sheet 4: 모듈별 집계 (auto COUNTIFS)
# ============================================================
def sheet_module_summary(wb):
    ws = wb.create_sheet('모듈별 집계')
    ws.sheet_properties.tabColor = SUMMARY_TAB

    headers = ['모듈', 'Critical', 'Major', 'Minor', '개선', '총건수', '해결완료', '해결률(%)']
    widths = [18, 12, 12, 12, 10, 12, 14, 14]
    apply_header(ws, headers, widths)

    for i, mod in enumerate(MODULES):
        r = i + 2
        ws.cell(row=r, column=1, value=mod).font = BODY_FONT
        # Critical/Major/Minor/개선 — COUNTIFS by module + severity
        for j, sev in enumerate(SEVERITIES, start=2):
            ws.cell(row=r, column=j,
                    value=f'=COUNTIFS(\'피드백 로그\'!B:B,"{mod}",\'피드백 로그\'!G:G,"{sev}")').font = BODY_FONT
        # 총건수
        ws.cell(row=r, column=6, value=f'=SUM(B{r}:E{r})').font = BODY_FONT
        # 해결완료
        ws.cell(row=r, column=7,
                value=f'=COUNTIFS(\'피드백 로그\'!B:B,"{mod}",\'피드백 로그\'!K:K,"종료")').font = BODY_FONT
        # 해결률(%)
        ws.cell(row=r, column=8, value=f'=IFERROR(G{r}/F{r}*100,0)').number_format = '0.0'
        for col in range(1, 9):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).alignment = ALIGN_TOP_WRAP

    # 합계 row
    total_row = len(MODULES) + 2
    ws.cell(row=total_row, column=1, value='합계').font = HEADER_FONT
    ws.cell(row=total_row, column=1).fill = HEADER_FILL
    for col_idx in range(2, 8):
        col_letter = get_column_letter(col_idx)
        ws.cell(row=total_row, column=col_idx,
                value=f'=SUM({col_letter}2:{col_letter}{total_row - 1})').font = HEADER_FONT
        ws.cell(row=total_row, column=col_idx).fill = HEADER_FILL
    ws.cell(row=total_row, column=8,
            value=f'=IFERROR(G{total_row}/F{total_row}*100,0)').number_format = '0.0'
    ws.cell(row=total_row, column=8).font = HEADER_FONT
    ws.cell(row=total_row, column=8).fill = HEADER_FILL
    for col in range(1, 9):
        ws.cell(row=total_row, column=col).border = BORDER


# ============================================================
# Sheet 5: 심각도별 집계 (auto COUNTIFS)
# ============================================================
def sheet_severity_summary(wb):
    ws = wb.create_sheet('심각도별 집계')
    ws.sheet_properties.tabColor = SUMMARY_TAB

    headers = ['심각도'] + STATUSES + ['총건수']
    widths = [12] + [12] * len(STATUSES) + [12]
    apply_header(ws, headers, widths)

    for i, sev in enumerate(SEVERITIES):
        r = i + 2
        ws.cell(row=r, column=1, value=sev).font = BODY_FONT
        for j, status in enumerate(STATUSES, start=2):
            ws.cell(row=r, column=j,
                    value=f'=COUNTIFS(\'피드백 로그\'!G:G,"{sev}",\'피드백 로그\'!K:K,"{status}")').font = BODY_FONT
        # 총건수
        end_col = get_column_letter(1 + len(STATUSES))
        ws.cell(row=r, column=2 + len(STATUSES),
                value=f'=SUM(B{r}:{end_col}{r})').font = BODY_FONT
        for col in range(1, 3 + len(STATUSES)):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).alignment = ALIGN_TOP_WRAP

    # 합계
    total_row = len(SEVERITIES) + 2
    ws.cell(row=total_row, column=1, value='합계').font = HEADER_FONT
    ws.cell(row=total_row, column=1).fill = HEADER_FILL
    for col_idx in range(2, 3 + len(STATUSES)):
        col_letter = get_column_letter(col_idx)
        ws.cell(row=total_row, column=col_idx,
                value=f'=SUM({col_letter}2:{col_letter}{total_row - 1})').font = HEADER_FONT
        ws.cell(row=total_row, column=col_idx).fill = HEADER_FILL
    for col in range(1, 3 + len(STATUSES)):
        ws.cell(row=total_row, column=col).border = BORDER


# ============================================================
# Sheet 6: UAT 판정
# ============================================================
def sheet_judgment(wb):
    ws = wb.create_sheet('UAT 판정')
    ws.sheet_properties.tabColor = JUDGMENT_TAB

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16

    ws['A1'] = 'CTR HR Hub UAT 최종 판정'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:E1')

    ws['A3'] = '통과 기준 체크리스트'
    ws['A3'].font = SECTION_FONT

    headers = ['#', '기준', '목표', '실제', '충족여부']
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = ALIGN_CENTER
        c.border = BORDER

    criteria = [
        (1, 'Critical 미해결 건수', '0건',
         '=COUNTIFS(\'피드백 로그\'!G:G,"Critical",\'피드백 로그\'!K:K,"<>종료")'),
        (2, 'Major 미해결 — 핵심 모듈 (인사/출퇴근/휴가/휴직/급여/결재·승인/설정)', '0건',
         '=COUNTIFS(\'피드백 로그\'!G:G,"Major",\'피드백 로그\'!K:K,"<>종료",\'피드백 로그\'!P:P,"차단")'),
        (3, 'Major 해결률 (전체)', '>=80%',
         '=IFERROR(COUNTIFS(\'피드백 로그\'!G:G,"Major",\'피드백 로그\'!K:K,"종료")/COUNTIF(\'피드백 로그\'!G:G,"Major")*100,0)'),
        (4, '핵심 시나리오 통과율', '>=95%', '(테스트북 집계 시트에서 확인)'),
        (5, '체크리스트 통과율', '>=90%', '(테스트북 집계 시트에서 확인)'),
        (6, '보안/데이터 이슈 미해결', '0건',
         '=COUNTIFS(\'피드백 로그\'!B:B,"규정준수",\'피드백 로그\'!K:K,"<>종료")'),
        (7, 'i18n 누락 0건', '0건',
         '=COUNTIFS(\'피드백 로그\'!D:D,"*i18n*",\'피드백 로그\'!K:K,"<>종료")+COUNTIFS(\'피드백 로그\'!D:D,"*번역*",\'피드백 로그\'!K:K,"<>종료")'),
    ]
    for i, (num, name, target, actual) in enumerate(criteria):
        r = 5 + i
        ws.cell(row=r, column=1, value=num).font = BODY_FONT
        ws.cell(row=r, column=2, value=name).font = BODY_FONT
        ws.cell(row=r, column=3, value=target).font = BODY_FONT
        ws.cell(row=r, column=4, value=actual).font = BODY_FONT
        ws.cell(row=r, column=5).font = BODY_FONT  # 충족여부 (수동)
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).alignment = ALIGN_TOP_WRAP

    dv_meet = DataValidation(type='list', formula1='"충족,미충족,N/A"', allow_blank=True)
    dv_meet.add(f'E5:E{4 + len(criteria)}')
    ws.add_data_validation(dv_meet)

    final_row = 4 + len(criteria) + 3
    ws.cell(row=final_row, column=1, value='최종 판정').font = SECTION_FONT

    options = [
        ('전체 GO — 운영 배포 진행', '[ ]'),
        ('조건부 GO — 명시된 조건 해결 후 배포', '[ ]'),
        ('NO-GO — 재테스트 필요', '[ ]'),
    ]
    for i, (text, mark) in enumerate(options):
        r = final_row + 1 + i
        ws.cell(row=r, column=2, value=text).font = BODY_FONT
        ws.cell(row=r, column=3, value=mark).font = BODY_FONT
        for col in (2, 3):
            ws.cell(row=r, column=col).border = BORDER

    cond_row = final_row + 5
    ws.cell(row=cond_row, column=1, value='조건부 GO 시 조건').font = SECTION_FONT
    for i in range(1, 5):
        ws.cell(row=cond_row + i, column=2, value=f'{i}.').font = BODY_FONT
        ws.merge_cells(start_row=cond_row + i, start_column=2,
                       end_row=cond_row + i, end_column=5)

    sign_row = cond_row + 6
    ws.cell(row=sign_row, column=1, value='사인오프').font = SECTION_FONT
    ws.cell(row=sign_row + 1, column=1, value='역할').font = HEADER_FONT
    ws.cell(row=sign_row + 1, column=2, value='이름').font = HEADER_FONT
    ws.cell(row=sign_row + 1, column=3, value='서명').font = HEADER_FONT
    ws.cell(row=sign_row + 1, column=4, value='날짜').font = HEADER_FONT
    for col in (1, 2, 3, 4):
        ws.cell(row=sign_row + 1, column=col).fill = HEADER_FILL
        ws.cell(row=sign_row + 1, column=col).alignment = ALIGN_CENTER
        ws.cell(row=sign_row + 1, column=col).border = BORDER

    sign_roles = ['UAT 관리자', 'HR 담당자 1', 'HR 담당자 2', 'CEO 승인']
    for i, role in enumerate(sign_roles):
        r = sign_row + 2 + i
        ws.cell(row=r, column=1, value=role).font = BODY_FONT
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).alignment = ALIGN_TOP_WRAP
        ws.row_dimensions[r].height = 28


# ============================================================
# Main
# ============================================================
def main():
    wb = Workbook()
    wb.remove(wb.active)

    sheet_guide(wb)
    sheet_log(wb)
    sheet_triage(wb)
    sheet_module_summary(wb)
    sheet_severity_summary(wb)
    sheet_judgment(wb)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    size_kb = OUT_PATH.stat().st_size / 1024
    print(f'✓ Wrote {OUT_PATH.relative_to(REPO_ROOT)} ({size_kb:.1f} KB)')
    print(f'  Sheets: {len(wb.sheetnames)}')
    for sn in wb.sheetnames:
        print(f'    - {sn}')


if __name__ == '__main__':
    main()
