#!/usr/bin/env python3
"""
CTR HR Hub — 사무직 UAT 피드백 폼 v2 (비즈니스 표준 톤)

Run: python3 scripts/uat/build-employee-feedback-v2.py
Output: docs/uat/UAT_피드백_사무직_v2.xlsx

v1 (deprecated) 대비 변경:
- 톤: 친화 구어체 → 비즈니스 표준 (1인칭 대화체 제거)
- 시트명: "🔴 이건 잘 안돼요" → "문제 보고", "💡 이게 있으면 좋겠어요" → "개선 제안"
- 심각도: "🔴 일이 안 돼요" → "심각 (업무 불가)"
- 우선순위: "⭐⭐⭐ 매일 도움" → "높음 (매일 활용)"
- 폰트: 12~13pt → 11pt (사무 표준)
- 줌: 110% → 100%
- 이모지 절제 (시트 탭 색상으로 구분, 텍스트 이모지 제거)
- 스크린샷: 선택사항 → 권장 (헤더에 명시)
- 컬럼 7개 분리 유지 — "발생 현상" + "기대 결과" 분리 (버그 리포트 표준)
- 안내 시트: FAQ 톤 비즈니스 한국어로 재작성

유지 (v1에서 가치 있던 부분):
- 18 → 7 컬럼 단순화
- 4 → 3 심각도
- 모듈명 한글 친화 (휴가 신청·조회 등 25개)
- 이슈ID·담당자·결함경과일 등 분류 메타 제거 (관리자가 v2 피드백 로그로 매핑)
- 문제 보고 vs 개선 제안 시트 분리 (분류 부담 ↓)
- 50 행 빈 슬롯 (한 워크북에 약 100건 수용)
- 자동 COUNTIFS 집계 (관리자 시트)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[2]
OUT_PATH = REPO_ROOT / 'docs/uat/UAT_피드백_사무직_v2.xlsx'

# ============================================================
# Styles — business standard
# ============================================================
FONT_NAME = 'Malgun Gothic'
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color='FFFFFF')
TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True, color='1F4E78')
SECTION_FONT = Font(name=FONT_NAME, size=11, bold=True, color='1F4E78')
BODY_FONT = Font(name=FONT_NAME, size=11)
BOLD_FONT = Font(name=FONT_NAME, size=11, bold=True)

# Sheet tab colors (subtle, business)
TAB_GUIDE = '595959'        # dark gray
TAB_PROBLEM = 'A52A2A'      # subdued red
TAB_SUGGESTION = 'B8860B'   # dark goldenrod
TAB_ADMIN = '4B0082'        # indigo

# Header fills
FILL_HEADER = PatternFill('solid', fgColor='2E75B6')     # blue (consistent across sheets)
FILL_HEADER_ADMIN = PatternFill('solid', fgColor='4B0082')

# Body accents
SOFT_GRAY = PatternFill('solid', fgColor='F2F2F2')
SOFT_BLUE = PatternFill('solid', fgColor='DDEBF7')
SOFT_GREEN = PatternFill('solid', fgColor='E2EFDA')
SOFT_RED = PatternFill('solid', fgColor='FCE4D6')
CALLOUT_FILL = PatternFill('solid', fgColor='FFF2CC')

ALIGN_TOP_WRAP = Alignment(vertical='top', wrap_text=True)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT_CENTER = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ============================================================
# Dropdown values
# ============================================================
SCREENS = [
    '홈 / 알림',
    '나의 공간 (내 정보)',
    '휴가 신청·조회',
    '휴직 신청·조회',
    '출퇴근',
    '결재함 (받은 결재)',
    '내 결재 신청 내역',
    '급여명세서',
    '연말 정산',
    '내 목표·평가',
    '분기 리뷰',
    '인정·칭찬',
    '리포트·통계',
    '직원 관리 (HR)',
    '조직도·부서 관리 (HR)',
    '근태 관리 (HR)',
    '휴가 관리 (HR)',
    '급여 관리 (HR)',
    '결재 흐름 설정 (HR)',
    '채용 관리 (HR)',
    '평가 관리 (HR)',
    '보상 관리 (HR)',
    '온보딩·오프보딩 (HR)',
    '마스터데이터·설정 (HR)',
    '컴플라이언스 (HR)',
    '기타',
]

# 3-level severity
SEVERITY = [
    '심각 — 업무 진행 불가',
    '중요 — 우회 가능',
    '경미 — 사소한 표시 오류',
]

# 3-level priority
PRIORITY = [
    '높음 — 매일 활용 예상',
    '중간 — 주 1~2회 활용',
    '낮음 — 가끔 활용',
]


# ============================================================
# Helpers
# ============================================================
def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_header_row(ws, headers, fill=FILL_HEADER, row=1):
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = fill
        c.alignment = ALIGN_CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 32
    ws.freeze_panes = f'A{row + 1}'


def write_empty_rows(ws, n_rows, n_cols, start_row=2, height=60):
    for r in range(start_row, start_row + n_rows):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = ALIGN_TOP_WRAP
            cell.border = BORDER
        ws.row_dimensions[r].height = height


# ============================================================
# Sheet 1: 작성 안내
# ============================================================
def sheet_guide(wb):
    ws = wb.create_sheet('작성 안내')
    ws.sheet_properties.tabColor = TAB_GUIDE

    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 100

    ws['B1'] = 'UAT 피드백 작성 안내'
    ws['B1'].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    r = 3

    # 목적
    ws.cell(row=r, column=2, value='1. 목적').font = SECTION_FONT
    r += 1
    intro = [
        'CTR HR Hub UAT(사용자 수용 테스트) 진행 중 발견하신 문제와 개선 제안을 기록하는 문서입니다.',
        '본 폼에 입력하신 내용은 UAT 관리자가 매일 검토 후 개발팀에 정식 이슈로 전달합니다.',
    ]
    for line in intro:
        ws.cell(row=r, column=2, value=line).font = BODY_FONT
        ws.cell(row=r, column=2).alignment = ALIGN_LEFT_CENTER
        ws.row_dimensions[r].height = 22
        r += 1

    r += 1
    # 시트 안내
    ws.cell(row=r, column=2, value='2. 시트 구성').font = SECTION_FONT
    r += 1
    sheets_info = [
        '• 문제 보고 — 오류, 비정상 동작, 잘못된 결과 등을 보고합니다.',
        '• 개선 제안 — 신규 기능, 사용성 개선, 편의 기능 등을 제안합니다.',
        '• 집계 — UAT 관리자 전용입니다. 작성자는 입력하실 필요 없습니다.',
    ]
    for line in sheets_info:
        ws.cell(row=r, column=2, value=line).font = BODY_FONT
        ws.row_dimensions[r].height = 22
        r += 1

    r += 1
    # 컬럼 설명
    ws.cell(row=r, column=2, value='3. 컬럼 작성 방법').font = SECTION_FONT
    r += 1
    columns_info = [
        ('작성자', '본인 성함을 입력합니다.'),
        ('일자', '발견·작성 날짜 (YYYY-MM-DD).'),
        ('화면', '드롭다운에서 해당 화면을 선택합니다. 정확한 화면 명칭이 없으면 "기타"를 선택하고 비고에 보충 설명을 적어주세요.'),
        ('발생 현상 / 현재 아쉬운 점', '관찰된 사실을 구체적으로 기술합니다. 입력값, 클릭한 버튼, 표시된 결과 등을 시간 순서대로 적습니다.'),
        ('기대 결과 / 제안 사항', '시스템이 어떻게 동작해야 하는지, 또는 어떤 기능이 추가되어야 하는지를 기술합니다.'),
        ('심각도 / 우선순위', '드롭다운에서 한 가지를 선택합니다. 분류가 어려운 경우 보수적으로(높은 등급 쪽) 선택해주세요.'),
        ('스크린샷', '권장 사항입니다. 화면 깨짐·오류 메시지 등 시각 정보가 도움 되는 경우 PrintScreen 후 셀에 붙여넣기(Ctrl+V) 합니다. 개인정보·민감정보는 마스킹해주세요.'),
    ]
    for col_name, desc in columns_info:
        ws.cell(row=r, column=2, value=f'• {col_name}: {desc}').font = BODY_FONT
        ws.cell(row=r, column=2).alignment = ALIGN_LEFT_CENTER
        ws.row_dimensions[r].height = 32
        r += 1

    r += 1
    # 심각도 / 우선순위 정의
    ws.cell(row=r, column=2, value='4. 심각도·우선순위 정의').font = SECTION_FONT
    r += 1
    severity_defs = [
        ('심각도 (문제 보고)', None),
        ('  · 심각 — 업무 진행 불가', '시스템이 정상 작동하지 않아 해당 업무를 수행할 수 없는 상태'),
        ('  · 중요 — 우회 가능', '비정상 동작이 있으나 다른 방법으로는 업무 수행이 가능한 상태'),
        ('  · 경미 — 사소한 표시 오류', 'UI 표시 오류, 오타, 정렬 불량 등 업무에 지장이 없는 사항'),
        ('', ''),
        ('우선순위 (개선 제안)', None),
        ('  · 높음 — 매일 활용 예상', '도입 시 일상 업무에서 매일 사용할 것으로 예상되는 기능'),
        ('  · 중간 — 주 1~2회 활용', '주기적으로 사용하는 기능 (주 1~2회 수준)'),
        ('  · 낮음 — 가끔 활용', '특정 시점·이벤트에만 사용하는 기능'),
    ]
    for label, desc in severity_defs:
        if desc is None:
            ws.cell(row=r, column=2, value=label).font = BOLD_FONT
            ws.row_dimensions[r].height = 22
        elif label == '':
            ws.row_dimensions[r].height = 8
        else:
            ws.cell(row=r, column=2, value=f'{label}: {desc}').font = BODY_FONT
            ws.cell(row=r, column=2).alignment = ALIGN_LEFT_CENTER
            ws.row_dimensions[r].height = 22
        r += 1

    r += 1
    # 좋은 예시
    ws.cell(row=r, column=2, value='5. 작성 예시 (권장)').font = SECTION_FONT
    r += 1

    examples_good = [
        '[문제 보고 예시]',
        '  화면: 휴가 신청·조회',
        '  발생 현상: 2026-05-25 09:00~13:00 오전 반차를 신청하고 매니저 승인을 완료했으나, 잔여 연차가 1일 차감되었습니다.',
        '  기대 결과: 반차이므로 0.5일만 차감되어야 합니다. 잔여 연차 표시는 14.5일이 되어야 합니다.',
        '  심각도: 중요 — 우회 가능 (수동으로 잔액 조정 가능)',
        '',
        '[개선 제안 예시]',
        '  화면: 휴가 신청·조회',
        '  현재 아쉬운 점: 휴가 신청 후 결재 진행 상태(현재 어느 단계, 누가 결재 대기 중)가 상세 페이지에 들어가야만 확인 가능합니다.',
        '  제안 사항: 신청 목록 화면에 결재자 성명과 진행 단계(예: 1차 승인 완료, 최종 승인 대기)를 함께 표시합니다.',
        '  우선순위: 높음 — 매일 활용 예상',
    ]
    for line in examples_good:
        cell = ws.cell(row=r, column=2, value=line)
        cell.font = BODY_FONT
        if line:
            cell.fill = SOFT_GREEN
        cell.alignment = ALIGN_LEFT_CENTER
        ws.row_dimensions[r].height = 22
        r += 1

    r += 1
    # 피해야 할 작성
    ws.cell(row=r, column=2, value='6. 피해야 할 작성 (재현 불가)').font = SECTION_FONT
    r += 1
    examples_bad = [
        '아래는 정보 부족으로 개발팀이 재현·수정할 수 없는 작성 예입니다:',
        '',
        '  ✗ 발생 현상: "휴가가 이상합니다."',
        '  ✗ 기대 결과: "정상적으로 작동하면 좋겠습니다."',
        '   → 어느 화면에서 어떤 동작이 어떻게 잘못되었는지 불명확합니다.',
        '',
        '  ✗ 발생 현상: "느립니다."',
        '   → "느리다"의 기준(N초), 어떤 액션, 어떤 조건인지 불명확합니다. "5건 조회 시 약 10초 대기" 식으로 정량 정보를 포함해주세요.',
    ]
    for line in examples_bad:
        cell = ws.cell(row=r, column=2, value=line)
        cell.font = BODY_FONT
        if line and line.startswith('  ✗'):
            cell.fill = SOFT_RED
        cell.alignment = ALIGN_LEFT_CENTER
        ws.row_dimensions[r].height = 22
        r += 1

    r += 1
    # 자주 묻는 질문
    ws.cell(row=r, column=2, value='7. 자주 묻는 질문').font = SECTION_FONT
    r += 1
    qa = [
        ('한 건당 어느 정도의 분량으로 작성해야 합니까?',
         '3~5줄이면 충분합니다. 화면명, 발생 현상, 기대 결과만 명확하면 됩니다.'),
        ('동일한 문제를 여러 번 발견했습니다.',
         '같은 사용자가 동일 건을 반복 입력할 필요는 없습니다. 비고 컬럼에 발생 횟수를 추가 기재해주세요.'),
        ('다른 사용자가 이미 보고한 것으로 보입니다.',
         '중복으로 보이더라도 별도 행으로 입력해주세요. UAT 관리자가 통합 처리합니다. 동일 건이 다수 보고되면 우선순위가 상승합니다.'),
        ('문제인지 개선인지 판단하기 어렵습니다.',
         '시스템이 의도대로 동작했으나 불편한 경우 → 개선 제안. 시스템이 의도와 다르게 동작한 경우 → 문제 보고.'),
        ('스크린샷에 개인정보가 포함되어 있습니다.',
         '주민번호, 급여, 평가 등급 등 민감정보는 사각형 등으로 마스킹 후 첨부해주세요.'),
    ]
    for q, a in qa:
        ws.cell(row=r, column=2, value=f'Q. {q}').font = BOLD_FONT
        ws.row_dimensions[r].height = 22
        r += 1
        ws.cell(row=r, column=2, value=f'A. {a}').font = BODY_FONT
        ws.cell(row=r, column=2).alignment = ALIGN_LEFT_CENTER
        ws.row_dimensions[r].height = 26
        r += 1
        r += 1

    # 마지막 안내
    cell = ws.cell(row=r, column=2,
                   value='문의 사항은 UAT 관리자에게 직접 연락해주십시오. 작성 형식이 완벽하지 않아도 무방합니다. '
                         '관찰하신 사실을 그대로 기록해주시는 것이 가장 중요합니다.')
    cell.font = BODY_FONT
    cell.fill = CALLOUT_FILL
    cell.alignment = ALIGN_LEFT_CENTER
    ws.row_dimensions[r].height = 40

    ws.sheet_view.zoomScale = 100


# ============================================================
# Sheet 2: 문제 보고
# ============================================================
def sheet_problem(wb):
    ws = wb.create_sheet('문제 보고')
    ws.sheet_properties.tabColor = TAB_PROBLEM

    headers = [
        '작성자',
        '발견 일자',
        '발생 화면',
        '발생 현상',
        '기대 결과',
        '심각도',
        '스크린샷 (권장)',
    ]
    widths = [12, 14, 24, 50, 40, 26, 22]
    set_col_widths(ws, widths)
    write_header_row(ws, headers)

    write_empty_rows(ws, 50, len(headers), start_row=2, height=60)

    dv_screen = DataValidation(type='list', formula1=f'"{",".join(SCREENS)}"', allow_blank=True)
    dv_screen.add('C2:C51')
    ws.add_data_validation(dv_screen)

    dv_sev = DataValidation(type='list', formula1=f'"{",".join(SEVERITY)}"', allow_blank=True)
    dv_sev.add('F2:F51')
    ws.add_data_validation(dv_sev)

    # Conditional formatting — severity color cues
    ws.conditional_formatting.add(
        'F2:F51',
        CellIsRule(operator='equal', formula=[f'"{SEVERITY[0]}"'],
                   fill=PatternFill('solid', fgColor='FFC7CE'),
                   font=Font(name=FONT_NAME, size=11, bold=True, color='9C0006')))
    ws.conditional_formatting.add(
        'F2:F51',
        CellIsRule(operator='equal', formula=[f'"{SEVERITY[1]}"'],
                   fill=PatternFill('solid', fgColor='FFEB9C'),
                   font=Font(name=FONT_NAME, size=11, color='9C5700')))
    ws.conditional_formatting.add(
        'F2:F51',
        CellIsRule(operator='equal', formula=[f'"{SEVERITY[2]}"'],
                   fill=PatternFill('solid', fgColor='E2EFDA'),
                   font=Font(name=FONT_NAME, size=11, color='375623')))

    ws.sheet_view.zoomScale = 100


# ============================================================
# Sheet 3: 개선 제안
# ============================================================
def sheet_suggestion(wb):
    ws = wb.create_sheet('개선 제안')
    ws.sheet_properties.tabColor = TAB_SUGGESTION

    headers = [
        '작성자',
        '작성 일자',
        '관련 화면',
        '현재 아쉬운 점',
        '제안 사항',
        '우선순위',
        '스크린샷 (권장)',
    ]
    widths = [12, 14, 24, 50, 40, 26, 22]
    set_col_widths(ws, widths)
    write_header_row(ws, headers)

    write_empty_rows(ws, 50, len(headers), start_row=2, height=60)

    dv_screen = DataValidation(type='list', formula1=f'"{",".join(SCREENS)}"', allow_blank=True)
    dv_screen.add('C2:C51')
    ws.add_data_validation(dv_screen)

    dv_pri = DataValidation(type='list', formula1=f'"{",".join(PRIORITY)}"', allow_blank=True)
    dv_pri.add('F2:F51')
    ws.add_data_validation(dv_pri)

    # Conditional formatting — priority (neutral business tones)
    ws.conditional_formatting.add(
        'F2:F51',
        CellIsRule(operator='equal', formula=[f'"{PRIORITY[0]}"'],
                   fill=PatternFill('solid', fgColor='DDEBF7'),
                   font=Font(name=FONT_NAME, size=11, bold=True, color='1F4E78')))
    ws.conditional_formatting.add(
        'F2:F51',
        CellIsRule(operator='equal', formula=[f'"{PRIORITY[1]}"'],
                   fill=PatternFill('solid', fgColor='F2F2F2'),
                   font=Font(name=FONT_NAME, size=11)))
    ws.conditional_formatting.add(
        'F2:F51',
        CellIsRule(operator='equal', formula=[f'"{PRIORITY[2]}"'],
                   fill=PatternFill('solid', fgColor='FAFAFA'),
                   font=Font(name=FONT_NAME, size=11, color='595959')))

    ws.sheet_view.zoomScale = 100


# ============================================================
# Sheet 4: 집계 (UAT 관리자 전용)
# ============================================================
def sheet_admin(wb):
    ws = wb.create_sheet('집계')
    ws.sheet_properties.tabColor = TAB_ADMIN

    ws.column_dimensions['A'].width = 32
    for c in range(2, 7):
        ws.column_dimensions[get_column_letter(c)].width = 18

    ws['A1'] = '집계 (UAT 관리자 전용)'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = 28

    note = ws.cell(row=2, column=1,
                   value='작성자께서는 본 시트를 별도로 채우지 않으셔도 됩니다. 다른 시트 입력 시 자동 집계됩니다.')
    note.font = BODY_FONT
    note.fill = CALLOUT_FILL
    note.alignment = ALIGN_LEFT_CENTER
    ws.merge_cells('A2:F2')
    ws.row_dimensions[2].height = 22

    sheet_problem_name = "'문제 보고'"
    sheet_suggest_name = "'개선 제안'"

    # Section 1: 문제 보고 — 화면별
    r = 4
    ws.cell(row=r, column=1, value='1. 문제 보고 — 화면별 건수').font = SECTION_FONT
    r += 1

    headers = ['화면', '심각', '중요', '경미', '합계']
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = FILL_HEADER_ADMIN
        c.alignment = ALIGN_CENTER
        c.border = BORDER
    ws.row_dimensions[r].height = 24
    r += 1

    for screen in SCREENS:
        ws.cell(row=r, column=1, value=screen).font = BODY_FONT
        ws.cell(row=r, column=2,
                value=f'=COUNTIFS({sheet_problem_name}!C2:C51,A{r},{sheet_problem_name}!F2:F51,"{SEVERITY[0]}")').font = BODY_FONT
        ws.cell(row=r, column=3,
                value=f'=COUNTIFS({sheet_problem_name}!C2:C51,A{r},{sheet_problem_name}!F2:F51,"{SEVERITY[1]}")').font = BODY_FONT
        ws.cell(row=r, column=4,
                value=f'=COUNTIFS({sheet_problem_name}!C2:C51,A{r},{sheet_problem_name}!F2:F51,"{SEVERITY[2]}")').font = BODY_FONT
        ws.cell(row=r, column=5, value=f'=SUM(B{r}:D{r})').font = BODY_FONT
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = ALIGN_TOP_WRAP
        r += 1

    total_row = r
    ws.cell(row=total_row, column=1, value='합계').font = HEADER_FONT
    ws.cell(row=total_row, column=1).fill = FILL_HEADER_ADMIN
    for c in range(2, 6):
        col_letter = get_column_letter(c)
        cell = ws.cell(row=total_row, column=c,
                       value=f'=SUM({col_letter}{total_row - len(SCREENS)}:{col_letter}{total_row - 1})')
        cell.font = HEADER_FONT
        cell.fill = FILL_HEADER_ADMIN
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER
    ws.cell(row=total_row, column=1).border = BORDER
    ws.cell(row=total_row, column=1).alignment = ALIGN_CENTER

    # Section 2: 개선 제안 — 화면별
    r = total_row + 3
    ws.cell(row=r, column=1, value='2. 개선 제안 — 화면별 건수').font = SECTION_FONT
    r += 1

    headers = ['화면', '높음', '중간', '낮음', '합계']
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = FILL_HEADER_ADMIN
        c.alignment = ALIGN_CENTER
        c.border = BORDER
    ws.row_dimensions[r].height = 24
    r += 1

    for screen in SCREENS:
        ws.cell(row=r, column=1, value=screen).font = BODY_FONT
        ws.cell(row=r, column=2,
                value=f'=COUNTIFS({sheet_suggest_name}!C2:C51,A{r},{sheet_suggest_name}!F2:F51,"{PRIORITY[0]}")').font = BODY_FONT
        ws.cell(row=r, column=3,
                value=f'=COUNTIFS({sheet_suggest_name}!C2:C51,A{r},{sheet_suggest_name}!F2:F51,"{PRIORITY[1]}")').font = BODY_FONT
        ws.cell(row=r, column=4,
                value=f'=COUNTIFS({sheet_suggest_name}!C2:C51,A{r},{sheet_suggest_name}!F2:F51,"{PRIORITY[2]}")').font = BODY_FONT
        ws.cell(row=r, column=5, value=f'=SUM(B{r}:D{r})').font = BODY_FONT
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = ALIGN_TOP_WRAP
        r += 1

    total_row2 = r
    ws.cell(row=total_row2, column=1, value='합계').font = HEADER_FONT
    ws.cell(row=total_row2, column=1).fill = FILL_HEADER_ADMIN
    for c in range(2, 6):
        col_letter = get_column_letter(c)
        cell = ws.cell(row=total_row2, column=c,
                       value=f'=SUM({col_letter}{total_row2 - len(SCREENS)}:{col_letter}{total_row2 - 1})')
        cell.font = HEADER_FONT
        cell.fill = FILL_HEADER_ADMIN
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER
    ws.cell(row=total_row2, column=1).border = BORDER
    ws.cell(row=total_row2, column=1).alignment = ALIGN_CENTER

    # Section 3: 전체 요약
    r = total_row2 + 3
    ws.cell(row=r, column=1, value='3. 전체 요약').font = SECTION_FONT
    r += 1

    summary = [
        ('전체 문제 보고 건수', f'=COUNTA({sheet_problem_name}!A2:A51)'),
        ('  심각', f'=COUNTIF({sheet_problem_name}!F2:F51,"{SEVERITY[0]}")'),
        ('  중요', f'=COUNTIF({sheet_problem_name}!F2:F51,"{SEVERITY[1]}")'),
        ('  경미', f'=COUNTIF({sheet_problem_name}!F2:F51,"{SEVERITY[2]}")'),
        ('전체 개선 제안 건수', f'=COUNTA({sheet_suggest_name}!A2:A51)'),
        ('  높음', f'=COUNTIF({sheet_suggest_name}!F2:F51,"{PRIORITY[0]}")'),
        ('  중간', f'=COUNTIF({sheet_suggest_name}!F2:F51,"{PRIORITY[1]}")'),
        ('  낮음', f'=COUNTIF({sheet_suggest_name}!F2:F51,"{PRIORITY[2]}")'),
        ('참여 작성자 수 (문제 보고 기준)',
         f'=SUMPRODUCT((COUNTIF({sheet_problem_name}!A2:A51,{sheet_problem_name}!A2:A51)<>0)/COUNTIF({sheet_problem_name}!A2:A51,{sheet_problem_name}!A2:A51&""))'),
    ]
    for label, formula in summary:
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = BODY_FONT
        c1.alignment = Alignment(horizontal='left', vertical='center')
        c1.border = BORDER
        c2 = ws.cell(row=r, column=2, value=formula)
        c2.font = BOLD_FONT
        c2.alignment = ALIGN_CENTER
        c2.border = BORDER
        ws.row_dimensions[r].height = 22
        r += 1

    # Section 4: 관리자 매핑 가이드
    r += 2
    ws.cell(row=r, column=1, value='4. 관리자 매핑 가이드').font = SECTION_FONT
    r += 1
    mapping_note = (
        '본 폼은 사무직 작성자가 자유롭게 입력하는 1차 수집 채널입니다. UAT 관리자는 다음을 수행합니다.\n\n'
        '  1) 신규 입력 검토 (일 1회 권장)\n'
        '  2) 개발자용 UAT_피드백_결과_v2.xlsx 피드백 로그에 FB-XXX 이슈로 이관:\n'
        '     - 화면 → 정확한 시스템 모듈 매핑\n'
        '     - 심각/중요/경미 → Critical/Major/Minor 매핑 (필요 시 격상·격하)\n'
        '     - 우선순위 → 개선(enhancement) 항목으로 분류\n'
        '  3) 중복 건은 본 폼의 비고에 "FB-XXX 동일" 표기 후 통합\n'
        '  4) 매주 통계는 본 시트 자동 집계 + 개발자 v2의 모듈별·심각도별 집계 시트를 함께 활용'
    )
    cell = ws.cell(row=r, column=1, value=mapping_note)
    cell.font = BODY_FONT
    cell.fill = CALLOUT_FILL
    cell.alignment = ALIGN_LEFT_CENTER
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 160

    ws.sheet_view.zoomScale = 100


# ============================================================
# Main
# ============================================================
def main():
    wb = Workbook()
    wb.remove(wb.active)

    sheet_guide(wb)
    sheet_problem(wb)
    sheet_suggestion(wb)
    sheet_admin(wb)

    wb.active = 0

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    size_kb = OUT_PATH.stat().st_size / 1024
    print(f'✓ Wrote {OUT_PATH.relative_to(REPO_ROOT)} ({size_kb:.1f} KB)')
    print(f'  Sheets: {len(wb.sheetnames)}')
    for sn in wb.sheetnames:
        print(f'    - {sn}')


if __name__ == '__main__':
    main()
