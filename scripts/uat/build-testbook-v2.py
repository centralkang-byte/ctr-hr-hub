#!/usr/bin/env python3
"""
CTR HR Hub UAT 테스트북 v2 generator
Run: python3 scripts/uat/build-testbook-v2.py
Output: docs/uat/UAT_테스트북_v2.xlsx

v2 changes vs v1:
- Test account count corrected to 8 (was 9 in v1)
- 휴직(LOA) split into its own sheet (was embedded in 휴가관리)
- 신규 시트 추가: V2 홈, 알림
- 신규 시나리오 추가: Compensation Letter PDF + Off-Cycle Comp 3 paths + Total Rewards +
  GoalRevision + QuarterlyReview 듀얼제출 + Calibration DnD + dept_head 결재 +
  multi-role canApprove + 인수인계 4 게이트 + do-not-rehire + 지정연차 자동 차감 +
  EDGE-026~030 personas
- 모듈별 집계 시트에 휴직 추가, COUNTIF/COUNTIFS 수식 자동 적용
- 직급 체계 정정 (국내 E1/S1/L1/L2 + 해외 L1~L5/S1/E1)
- 연차촉진 시나리오를 "수동 통보 점검"으로 정정 (cron 미동작 명시)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[2]
OUT_PATH = REPO_ROOT / 'docs/uat/UAT_테스트북_v2.xlsx'

# ============================================================
# Styles
# ============================================================
FONT_NAME = 'Malgun Gothic'
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
HEADER_FILL = PatternFill('solid', fgColor='2E75B6')
HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color='FFFFFF')
SECTION_FILL = PatternFill('solid', fgColor='DDEBF7')
SECTION_FONT = Font(name=FONT_NAME, size=11, bold=True, color='1F4E78')
TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True, color='1F4E78')
BODY_FONT = Font(name=FONT_NAME, size=11)
NEW_FILL = PatternFill('solid', fgColor='FFF2CC')  # 신규 시나리오 강조
ALIGN_TOP_WRAP = Alignment(vertical='top', wrap_text=True)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Tab colors per category
SCENARIO_TAB = '4472C4'   # blue (시나리오)
CHECKLIST_TAB = '70AD47'  # green (체크리스트)
META_TAB = 'A6A6A6'       # gray (사용법/집계)
EDGE_TAB = 'ED7D31'       # orange (엣지케이스)

# ============================================================
# Scenario column schema
# ============================================================
SCENARIO_HEADERS = [
    'Step#', '비즈니스 목적', '테스터 역할', '전제조건', '테스트 데이터',
    '수행 액션', '기대되는 비즈니스 결과', '실제 결과', 'Pass/Fail/Block',
    '이슈ID', '비고'
]
SCENARIO_WIDTHS = [6, 22, 22, 25, 30, 40, 40, 25, 14, 10, 18]

CHECKLIST_HEADERS = [
    '#', '기능', '세부기능', '테스트 계정', '확인 (Y/N/NA)',
    '심각도', '테스터', '날짜', '비고'
]
CHECKLIST_WIDTHS = [5, 20, 50, 22, 14, 12, 12, 12, 22]


def apply_header(ws, headers, widths, row=1):
    """Apply header row styling."""
    for col_idx, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = f'A{row + 1}'


def write_scenario_rows(ws, rows, start_row=2, mark_new=None):
    """Write scenario data rows. mark_new = set of step numbers to highlight as new."""
    mark_new = mark_new or set()
    for i, row in enumerate(rows):
        r = start_row + i
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = ALIGN_TOP_WRAP
            cell.border = BORDER
            if i + 1 in mark_new and col_idx <= 7:
                cell.fill = NEW_FILL
        # Add empty cells for 실제결과/Pass-Fail/이슈ID/비고
        for col_idx in range(len(row) + 1, len(SCENARIO_HEADERS) + 1):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = BORDER
            cell.alignment = ALIGN_TOP_WRAP
        ws.row_dimensions[r].height = 60


def write_checklist_rows(ws, rows, start_row=2):
    for i, row in enumerate(rows):
        r = start_row + i
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = ALIGN_TOP_WRAP
            cell.border = BORDER
        for col_idx in range(len(row) + 1, len(CHECKLIST_HEADERS) + 1):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = BORDER
        ws.row_dimensions[r].height = 24


def add_dropdown(ws, range_str, values):
    """Add a dropdown data validation."""
    dv = DataValidation(type='list', formula1=f'"{",".join(values)}"', allow_blank=True)
    dv.add(range_str)
    ws.add_data_validation(dv)


def add_conditional_pass_fail(ws, range_str):
    """Pass=green, Fail=red, Block=yellow."""
    ws.conditional_formatting.add(
        range_str,
        CellIsRule(operator='equal', formula=['"Pass"'],
                   fill=PatternFill('solid', fgColor='C6EFCE'),
                   font=Font(name=FONT_NAME, size=11, color='006100')))
    ws.conditional_formatting.add(
        range_str,
        CellIsRule(operator='equal', formula=['"Fail"'],
                   fill=PatternFill('solid', fgColor='FFC7CE'),
                   font=Font(name=FONT_NAME, size=11, color='9C0006')))
    ws.conditional_formatting.add(
        range_str,
        CellIsRule(operator='equal', formula=['"Block"'],
                   fill=PatternFill('solid', fgColor='FFEB9C'),
                   font=Font(name=FONT_NAME, size=11, color='9C5700')))


# ============================================================
# Sheet content
# ============================================================

def sheet_usage(wb):
    """사용법 sheet — corrected from v1."""
    ws = wb.create_sheet('사용법')
    ws.sheet_properties.tabColor = META_TAB

    rows = [
        ('CTR HR Hub UAT 테스트북 v2 사용법', None),
        ('', None),
        ('1. 문서 구성', None),
        ('이 테스트북은 두 가지 포맷으로 구성되어 있습니다.', None),
        ('  - 시나리오 (파란 탭): 단계별 비즈니스 플로우를 따라가며 검증합니다. 한 시나리오 안에서 여러 역할이 참여합니다.', None),
        ('  - 체크리스트 (초록 탭): 기능 단위로 빠르게 존재 여부와 동작을 확인합니다.', None),
        ('  - 신규 기능은 노란색 배경으로 강조됩니다.', None),
        ('', None),
        ('2. 시나리오 시트 컬럼 설명', None),
        ('  A. Step#  — 순서 번호', None),
        ('  B. 비즈니스 목적  — 이 단계가 왜 중요한지 (비즈니스 관점)', None),
        ('  C. 테스터 역할  — 어떤 계정으로 로그인해야 하는지', None),
        ('  D. 전제조건  — 이 단계 실행 전에 완료되어야 할 것', None),
        ('  E. 테스트 데이터  — 입력할 구체적인 데이터', None),
        ('  F. 수행 액션  — 실제로 할 행동 (한국어 단계별 설명)', None),
        ('  G. 기대되는 비즈니스 결과  — 정상일 때 화면에 보여야 할 것', None),
        ('  H. 실제 결과  — 테스터가 직접 기록', None),
        ('  I. Pass/Fail/Block  — 드롭다운으로 선택', None),
        ('  J. 이슈ID  — Fail/Block 시 이슈 추적 번호 (예: FB-007)', None),
        ('  K. 비고  — 추가 메모', None),
        ('', None),
        ('3. 체크리스트 시트 컬럼 설명', None),
        ('  A. #  — 순서 번호', None),
        ('  B. 기능  — 대분류 기능명', None),
        ('  C. 세부기능  — 구체적 페이지 또는 기능', None),
        ('  D. 테스트 계정  — 확인에 사용할 계정', None),
        ('  E. 확인  — Y(확인) / N(미확인) / NA(해당없음)', None),
        ('  F. 심각도  — N인 경우만: Critical / Major / Minor / 개선', None),
        ('  G. 테스터  — 확인한 사람', None),
        ('  H. 날짜  — 확인 날짜', None),
        ('  I. 비고  — 추가 메모', None),
        ('', None),
        ('4. 색상 범례', None),
        ('  Pass (녹색)  — 테스트 통과, 기대 결과와 일치', None),
        ('  Fail (빨간색)  — 테스트 실패, 기대 결과와 불일치 (이슈ID 필수 기록)', None),
        ('  Block (노란색)  — 차단, 선행 조건 미충족으로 테스트 불가', None),
        ('  N/A (회색)  — 해당 없음, 현재 환경에서 테스트 불가', None),
        ('  노란 배경 행  — v1 대비 신규 추가된 기능 (Sessions 168~217 추가분)', None),
        ('', None),
        ('5. 테스트 계정 (9개)', None),
        ('  대조영 — super@ctr.co.kr — SUPER_ADMIN — CTR-HOLD', None),
        ('  한지영 — hr@ctr.co.kr — HR_ADMIN — CTR', None),
        ('  陈美玲 — hr@ctr-cn.com — HR_ADMIN — CTR-CN', None),
        ('  박준혁 — manager@ctr.co.kr — MANAGER — CTR (QA-TEAM-A)', None),
        ('  김서연 — manager2@ctr.co.kr — MANAGER — CTR (QA-TEAM-B)', None),
        ('  이민준 — employee-a@ctr.co.kr — EMPLOYEE — CTR (QA-TEAM-A, → 박준혁)', None),
        ('  정다은 — employee-b@ctr.co.kr — EMPLOYEE — CTR (QA-TEAM-A, → 박준혁)', None),
        ('  송현우 — employee-c@ctr.co.kr — EMPLOYEE — CTR (QA-TEAM-B, → 김서연)', None),
        ('  강대표 — executive@ctr.co.kr — EXECUTIVE — CTR (E1)', None),
        ('', None),
        ('6. 직급 체계 (Session 167 전환 반영)', None),
        ('  국내 (4단계): E1 (임원), S1 (수석), L2 (책임), L1 (선임/사원)', None),
        ('  해외 (7단계): L1~L5, S1, E1', None),
        ('', None),
        ('7. 주의사항', None),
        ('  - 이 문서는 E2E 자동 테스트와 다릅니다 — 비즈니스 결과를 확인하세요.', None),
        ('  - 시나리오 시트는 반드시 Step 순서대로 실행하세요 (선행 단계의 결과가 다음 단계의 전제조건입니다).', None),
        ('  - 체크리스트 시트는 순서 무관하게 독립적으로 확인할 수 있습니다.', None),
        ('  - UAT는 staging 환경 전용입니다 (운영 환경 prod는 SSO-only — Session 212 이후).', None),
        ('  - 테스트 계정 비밀번호 불필요: NEXT_PUBLIC_SHOW_TEST_ACCOUNTS=true 활성화된 staging URL에 카드 클릭만으로 로그인.', None),
        ('  - 집계 시트에서 전체 진행률을 확인할 수 있습니다 (COUNTIFS 수식 자동 집계).', None),
    ]

    ws.column_dimensions['A'].width = 100
    ws.cell(row=1, column=1, value=rows[0][0]).font = TITLE_FONT
    for i, (text, _) in enumerate(rows[1:], start=2):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = BODY_FONT
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        if text.startswith(('1.', '2.', '3.', '4.', '5.', '6.', '7.')):
            cell.font = SECTION_FONT


# ============================================================
# 인사관리 (Employee + Org)
# ============================================================
def sheet_employee(wb):
    ws = wb.create_sheet('인사관리')
    ws.sheet_properties.tabColor = SCENARIO_TAB
    apply_header(ws, SCENARIO_HEADERS, SCENARIO_WIDTHS)

    rows = [
        # (Step, 비즈니스 목적, 역할, 전제, 데이터, 액션, 기대결과)
        (1, '직원 목록 조회', 'HR_ADMIN (hr@ctr.co.kr)', '로그인 완료', '',
         '좌측 메뉴 > 인사관리 > 직원 관리 클릭',
         '직원 목록 페이지가 로드되고 CTR 법인 소속 직원이 표시됨'),
        (2, '직원 목록 필터링', 'HR_ADMIN', 'Step 1 완료', '부서: QA-TEAM-A',
         '필터 영역에서 부서를 "QA-TEAM-A"로 선택',
         '해당 부서 소속 직원(박준혁/이민준/정다은)만 표시됨'),
        (3, '직원 검색 (이름)', 'HR_ADMIN', 'Step 1 완료', '검색어: 이민준',
         '검색창에 "이민준" 입력', '이민준 직원이 검색 결과에 표시됨'),
        (4, '직원 검색 (사번)', 'HR_ADMIN', 'Step 1 완료', '검색어: EDGE-001',
         '검색창에 "EDGE-001" 입력', '김수습(EDGE-001) 직원이 검색 결과에 표시됨'),
        (5, '신규 직원 등록 시작', 'HR_ADMIN', 'Step 1 완료', '',
         '직원 목록 페이지에서 "신규 등록" 버튼 클릭', '신규 직원 등록 폼이 표시됨'),
        (6, '신규 직원 정보 입력', 'HR_ADMIN', 'Step 5 완료',
         '이름: 테스트직원, 부서: QA-TEAM-A, 직급: L1, 입사일: 2026-05-15',
         '필수 필드(이름, 부서, 직급, 입사일, 고용형태) 입력 후 저장',
         '직원이 성공적으로 등록되고 상세 페이지로 이동'),
        (7, '등록된 직원 상세 조회', 'HR_ADMIN', 'Step 6 완료', '',
         '등록된 직원 상세 페이지에서 기본정보 탭 확인',
         '입력한 이름, 부서, 직급, 입사일이 정확히 표시됨'),
        (8, '직원 정보 수정', 'HR_ADMIN', 'Step 7 완료', '변경할 연락처: 010-9999-0000',
         '수정 버튼 클릭 > 연락처 변경 > 저장', '연락처가 변경된 값으로 표시됨'),
        (9, '인사 이력 확인', 'HR_ADMIN', 'Step 7 완료', '',
         '직원 상세 > 발령 이력 탭 클릭',
         '해당 직원의 발령/이동 이력이 시간순으로 표시됨'),
        (10, '조직도 조회', 'HR_ADMIN', '로그인 완료', '',
         '좌측 메뉴 > 인사관리 > 조직 관리 클릭',
         '조직도가 트리 구조로 표시되고 보고 라인이 정확함'),
        (11, '디렉토리 조회', 'HR_ADMIN', '로그인 완료', '',
         '좌측 메뉴 > 디렉토리 클릭', '전체 직원 디렉토리가 카드/리스트 형태로 표시됨'),
        (12, '근로계약서 확인', 'HR_ADMIN', 'Step 7 완료', '',
         '직원 상세 > 계약 탭 클릭', '근로계약서 정보가 표시됨 (계약 유형, 기간 등)'),
        (13, '수습 직원 조회', 'HR_ADMIN', 'EDGE-001 시드 데이터 존재', '김수습 (EDGE-001)',
         '직원 목록에서 수습 상태 필터 적용',
         '김수습이 수습 배지와 함께 표시되고 수습 만료일(2026-06-01)이 확인됨'),
        (14, '수습 만료 직원 경고', 'HR_ADMIN', 'EDGE-002 시드 데이터 존재', '박만료 (EDGE-002)',
         '직원 목록에서 박만료 검색',
         '수습 기간 만료 경고가 표시됨 (만료일 2026-01-01 이미 지남)'),
        (15, '계약직 만료 예정자 확인', 'HR_ADMIN', 'EDGE-003 시드 데이터 존재', '이계약 (EDGE-003)',
         '직원 목록 또는 대시보드에서 계약 만료 예정자 확인',
         '이계약의 계약 만료 임박(D-30) 알림 또는 배지가 표시됨'),
        (16, '계약 만료 직원 상태', 'HR_ADMIN', 'EDGE-004 시드 데이터 존재', '최만기 (EDGE-004)',
         '직원 목록에서 최만기 검색', '계약 만료 상태가 명확히 표시됨'),
        (17, '휴직자 목록 확인', 'HR_ADMIN', 'EDGE-005 시드 데이터 존재', '정육아 (EDGE-005)',
         '직원 목록에서 휴직 상태 필터 또는 휴가/휴직 관리 메뉴',
         '정육아가 육아휴직 상태로 표시되고 출퇴근/급여 대상에서 제외됨'),
        (18, '퇴사 처리 시작', 'HR_ADMIN', 'Step 7에서 등록한 직원 존재',
         '퇴사일: 2026-05-31, 사유: 자발적 퇴사',
         '직원 상세 > 오프보딩 시작 > 퇴사일과 사유 입력',
         '오프보딩 프로세스가 시작되고 상태가 "퇴사 예정"으로 변경됨'),
        (19, '퇴직 완료 직원 확인', 'HR_ADMIN', 'EDGE-008 시드 데이터 존재', '유퇴직 (EDGE-008)',
         '직원 목록에서 유퇴직 검색 (퇴사자 포함 필터)',
         '유퇴직이 "퇴직" 상태로 표시되고 활성 직원 목록에서는 제외됨'),
        (20, 'EMPLOYEE 내 프로필 확인', 'EMPLOYEE (employee-a@ctr.co.kr)', '로그인 완료', '',
         '좌측 메뉴 > 나의 공간 > 내 프로필 클릭',
         '이민준의 프로필 정보(이름, 부서, 직급, 입사일)가 정확히 표시됨'),
        (21, 'EMPLOYEE 타인 정보 접근 차단', 'EMPLOYEE (employee-a@ctr.co.kr)',
         'Step 20 완료', 'URL: /employees/[다른직원ID]',
         '브라우저 주소창에 다른 직원의 상세 페이지 URL 직접 입력',
         '403 접근 거부 또는 권한 없음 메시지가 표시됨'),
        (22, 'MANAGER 팀원 목록 확인', 'MANAGER (manager@ctr.co.kr)', '로그인 완료', '',
         '좌측 메뉴 > 팀 관리 > 팀 현황 클릭',
         '박준혁 매니저의 직속 팀원(이민준, 정다은)만 표시됨'),
        (23, '부서 이동 발령', 'HR_ADMIN', 'Step 7 완료', '이동 대상: 테스트직원, 새 부서: QA-TEAM-B',
         '직원 상세 > 발령 처리 > 부서 이동 선택 > 새 부서 지정 > 확인',
         '발령 이력에 부서 이동이 기록되고 직원의 소속 부서가 변경됨'),
        (24, '직급 변경', 'HR_ADMIN', 'Step 23 완료', '새 직급: L2',
         '직원 상세 > 발령 처리 > 직급 변경 선택 > L2 지정 > 확인',
         '발령 이력에 직급 변경이 기록되고 직원의 직급이 L2로 변경됨'),
        (25, '겸직자 조회', 'HR_ADMIN', 'EDGE-009 시드 데이터 존재', '강겸직 (EDGE-009)',
         '직원 목록에서 강겸직 검색', '강겸직의 복수 법인 발령 정보가 모두 표시됨'),
        (26, '다직위자 조회', 'HR_ADMIN', 'EDGE-010 시드 데이터 존재', '임다직 (EDGE-010)',
         '직원 목록에서 임다직 검색 > 상세 페이지 진입',
         '임다직의 3개 직위가 모두 표시됨'),
        (27, '매니저 없는 직원 확인', 'HR_ADMIN', 'EDGE-011 시드 데이터 존재', '윤대표 (EDGE-011)',
         '직원 상세에서 윤대표의 보고 라인 확인',
         '상위 보고자가 없음(최상위)으로 표시되며 오류 없이 동작'),
        (28, '부서 미배정 직원', 'HR_ADMIN', 'EDGE-012 시드 데이터 존재', '신입사 (EDGE-012)',
         '직원 목록에서 신입사 검색',
         '부서 미배정 상태가 표시되며 오류 없이 목록에 나타남'),
        (29, '법인 간 전적 직원', 'HR_ADMIN', 'EDGE-013 시드 데이터 존재', '배전적 (EDGE-013)',
         '직원 상세에서 배전적의 이력 확인',
         '법인 전적(CTR→CTR-CN) 이력이 발령 이력에 기록됨'),
        (30, '일괄 발령 처리 (Bulk HR Movements)', 'HR_ADMIN', 'Step 1 완료',
         '대상: 복수 직원, 발령 유형: 부서 이동',
         '인사관리 > 일괄 발령 메뉴 진입 > 대상 직원 선택 > CSV 업로드 또는 수동 선택 > 발령 유형 지정 > 3단계 마법사',
         '선택한 직원들에 대해 일괄 발령이 처리되고 이력에 기록됨 (All-or-Nothing)'),
    ]

    write_scenario_rows(ws, rows)
    add_dropdown(ws, f'I2:I{len(rows) + 1}', ['Pass', 'Fail', 'Block', 'N/A'])
    add_conditional_pass_fail(ws, f'I2:I{len(rows) + 1}')
    return len(rows)


# ============================================================
# 출퇴근
# ============================================================
def sheet_attendance(wb):
    ws = wb.create_sheet('출퇴근')
    ws.sheet_properties.tabColor = SCENARIO_TAB
    apply_header(ws, SCENARIO_HEADERS, SCENARIO_WIDTHS)

    rows = [
        (1, '출근 기록', 'EMPLOYEE (employee-a@ctr.co.kr)', '로그인 완료, 근무일', '',
         '좌측 메뉴 > 나의 공간 > 출퇴근 > 출근 버튼 클릭',
         '출근 시각이 현재 시간으로 기록되고 상태가 "근무 중"으로 변경됨'),
        (2, '퇴근 기록', 'EMPLOYEE (employee-a@ctr.co.kr)', 'Step 1 완료 (출근 상태)', '',
         '나의 출퇴근 페이지에서 퇴근 버튼 클릭',
         '퇴근 시각이 기록되고 근무시간이 자동 계산됨'),
        (3, '출퇴근 기록 조회', 'EMPLOYEE (employee-a@ctr.co.kr)', 'Step 2 완료', '',
         '나의 출퇴근 페이지에서 이번 주 기록 확인',
         '금일 출근/퇴근 시각과 근무시간이 정확히 표시됨'),
        (4, 'MANAGER 팀 출퇴근 현황', 'MANAGER (manager@ctr.co.kr)', '팀원 출퇴근 기록 존재', '',
         '좌측 메뉴 > 팀 관리 > 팀 근태/휴가 클릭',
         '박준혁 매니저의 팀원(이민준, 정다은)의 출퇴근 현황이 표시됨'),
        (5, 'HR_ADMIN 출퇴근 관리', 'HR_ADMIN (hr@ctr.co.kr)', '로그인 완료', '',
         '좌측 메뉴 > 인사관리 > 근태 관리 클릭',
         '전체 직원의 출퇴근 현황이 관리자 뷰로 표시됨'),
        (6, '이상 출퇴근 감지 — 지각', 'HR_ADMIN', 'EDGE-023 시드 데이터 존재', '지각왕 (EDGE-023)',
         '출퇴근 관리에서 이상 항목 필터 적용',
         '지각왕의 빈번한 지각 기록이 이상 항목으로 감지됨'),
        (7, '이상 출퇴근 감지 — 결근', 'HR_ADMIN', 'Step 5 완료', '',
         '출퇴근 관리에서 결근 필터 적용', '미출근 직원이 결근으로 표시됨'),
        (8, '수동 출퇴근 조정', 'HR_ADMIN', 'Step 5 완료',
         '대상: 이민준, 조정 사유: 외근',
         '출퇴근 관리에서 특정 직원의 기록 선택 > 수정 > 사유 입력 > 저장',
         '수정된 출퇴근 기록이 반영되고 조정 이력이 기록됨'),
        (9, '교대근무 로스터 조회', 'HR_ADMIN', '교대근무 설정 존재', '',
         '근태 관리 > 교대근무 로스터 메뉴 클릭',
         '직원별 교대 배정이 표 형태로 표시됨 + 3교대 패턴 확인'),
        (10, '팀 마감 처리', 'MANAGER (manager@ctr.co.kr)', '월말', '',
         '팀 근태/휴가 > 마감 처리 버튼 클릭',
         '해당 기간의 팀 출퇴근이 마감되고 상태가 "마감완료"로 변경됨'),
        (11, '52시간 모니터링', 'HR_ADMIN', 'EDGE-024 시드 데이터 존재', '과로자 (EDGE-024)',
         '근태 관리 또는 대시보드에서 주 52시간 현황 확인',
         '과로자의 주간 근무시간이 52시간 경고와 함께 표시됨'),
        (12, '출퇴근 수정 요청', 'EMPLOYEE (employee-a@ctr.co.kr)',
         '출퇴근 수정 필요 상황', '수정 사유: 기기 오류로 미퇴근 기록',
         '나의 출퇴근 > 수정 요청 > 사유 입력 > 제출',
         '출퇴근 수정 요청이 결재함으로 전송됨'),
        (13, '출퇴근 수정 결재 처리', 'MANAGER (manager@ctr.co.kr)', 'Step 12 완료', '',
         '결재 > 결재함 > 출퇴근 수정 요청 확인 > 승인',
         '출퇴근 수정이 승인되고 기록이 반영됨'),
        (14, '초과근무 기록 확인', 'EMPLOYEE (employee-a@ctr.co.kr)', '초과근무 기록 존재', '',
         '나의 출퇴근에서 초과근무 탭 또는 영역 확인',
         '초과근무 시간이 별도로 집계되어 표시됨'),
        (15, '월간 출퇴근 리포트', 'HR_ADMIN', '한 달 이상의 출퇴근 데이터 존재', '',
         '근태 관리 > 리포트/통계 영역 확인',
         '월간 출퇴근 통계(총 근무일, 지각/결근 건수 등)가 표시됨'),
        (16, '휴직자 출퇴근 제외 확인', 'HR_ADMIN', 'EDGE-005 시드 데이터 존재', '정육아 (EDGE-005)',
         '근태 관리에서 정육아 검색',
         '정육아(육아휴직)가 출퇴근 대상에서 제외되어 있음'),
        (17, '퇴사 예정자 출퇴근', 'HR_ADMIN', 'EDGE-007 시드 데이터 존재', '오퇴사 (EDGE-007)',
         '근태 관리에서 오퇴사 확인', '오퇴사는 퇴사일까지 출퇴근 대상에 포함됨'),
        (18, 'GPS 펀치 (모바일)', 'EMPLOYEE (employee-a@ctr.co.kr)',
         '모바일 환경', '위치: 회사 주소',
         '모바일에서 출퇴근 화면 진입 > GPS 출근 버튼 탭',
         'GPS 위치가 허용 반경 내일 때 출근이 정상 기록됨 (반경 외는 거부)'),
        (19, '연장/야간/휴일 근무 구분', 'HR_ADMIN', '초과근무 데이터 존재', '',
         '근태 관리에서 초과근무 유형별 조회',
         '연장근무, 야간근무, 휴일근무가 구분되어 집계됨'),
        (20, '근태 이상 일괄 처리', 'HR_ADMIN', '복수 이상 출퇴근 건 존재', '',
         '근태 관리 > 이상 항목 > 일괄 처리',
         '선택한 이상 항목들이 일괄로 처리됨'),
        (21, '터미널/GPS 관리', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 출퇴근 > 터미널 관리 확인',
         '출퇴근 단말기/GPS 설정 화면이 표시됨'),
        (22, '출퇴근 데이터 가져오기', 'HR_ADMIN', '로그인 완료', '',
         '근태 관리 > 데이터 가져오기 메뉴 확인',
         '외부 출퇴근 데이터 임포트 기능이 제공됨'),
    ]
    write_scenario_rows(ws, rows)
    add_dropdown(ws, f'I2:I{len(rows) + 1}', ['Pass', 'Fail', 'Block', 'N/A'])
    add_conditional_pass_fail(ws, f'I2:I{len(rows) + 1}')
    return len(rows)


# ============================================================
# 휴가관리 (LOA removed - separate sheet)
# ============================================================
def sheet_leave(wb):
    ws = wb.create_sheet('휴가관리')
    ws.sheet_properties.tabColor = SCENARIO_TAB
    apply_header(ws, SCENARIO_HEADERS, SCENARIO_WIDTHS)

    rows = [
        (1, '연차 잔액 확인', 'EMPLOYEE (employee-a@ctr.co.kr)', '로그인 완료',
         '예상 잔액: 15일',
         '좌측 메뉴 > 나의 공간 > 휴가 신청 클릭',
         '이민준의 연차 잔액이 15일로 표시됨 (총/사용/잔여)'),
        (2, '연차 신청', 'EMPLOYEE (employee-a@ctr.co.kr)', 'Step 1 완료',
         '신청일: 2026-05-20~21 (2일), 사유: 개인 사유',
         '내 휴가 > 휴가 신청 > 카테고리 "연차" 선택 > 날짜 선택 > 사유 입력 > 제출',
         '휴가 신청이 접수되고 상태가 "승인 대기"로 표시됨'),
        (3, '신청 후 잔액 잠정 차감', 'EMPLOYEE (employee-a@ctr.co.kr)', 'Step 2 완료', '',
         '내 휴가 페이지에서 잔액 재확인',
         '가용 일수가 13일(승인 대기 포함하여 잠정 차감)로 표시됨'),
        (4, 'MANAGER 결재함 휴가 확인', 'MANAGER (manager@ctr.co.kr)', 'Step 2 완료', '',
         '좌측 메뉴 > 결재 > 결재함 클릭',
         '이민준의 연차 신청이 대기 건으로 표시됨'),
        (5, '휴가 승인', 'MANAGER (manager@ctr.co.kr)', 'Step 4 완료', '',
         '결재함에서 이민준의 연차 신청 선택 > 승인 버튼 클릭',
         '승인 완료되고 상태가 "승인"으로 변경됨'),
        (6, '승인 후 직원 잔액 확정', 'EMPLOYEE (employee-a@ctr.co.kr)', 'Step 5 완료', '',
         '내 휴가 페이지에서 잔액 재확인',
         '잔여 연차가 13일로 확정 표시됨 (승인 완료)'),
        (7, '반차 신청 (0.5일)', 'EMPLOYEE (employee-a@ctr.co.kr)', 'Step 6 완료',
         '신청일: 2026-05-22 오전 반차',
         '내 휴가 > 휴가 신청 > 반차(오전) 선택 > 제출',
         '반차 0.5일이 신청되고 잔액이 12.5일로 표시됨 (해당 휴가 유형이 반차 허용 시)'),
        (8, '병가 신청 (증빙 첨부)', 'EMPLOYEE (employee-a@ctr.co.kr)', 'Step 6 완료',
         '병가 1일, 진단서 PDF 첨부',
         '내 휴가 > 휴가 신청 > 카테고리 "건강" > 병가 선택 > 증빙 첨부 > 제출',
         '병가가 신청되고 연차 잔액과는 별도로 처리됨 (증빙 필수 정책 적용)'),
        (9, '휴가 반려 + 사유', 'MANAGER (manager@ctr.co.kr)', '직원 휴가 신청 대기 건 존재',
         '반려 사유: 업무 일정 충돌',
         '결재함 > 휴가 신청 선택 > 반려 > 사유 입력 > 확인',
         '반려 처리되고 직원에게 반려 사유가 전달됨'),
        (10, '반려 후 직원 확인', 'EMPLOYEE (employee-a@ctr.co.kr)', 'Step 9 완료', '',
         '내 휴가에서 반려된 신청 확인',
         '반려된 신청이 "반려" 상태로 표시되고 사유가 확인 가능 + 잔액 복구'),
        (11, 'HR_ADMIN 휴가 관리', 'HR_ADMIN (hr@ctr.co.kr)', '로그인 완료', '',
         '좌측 메뉴 > 인사관리 > 휴가/휴직 관리 클릭',
         '전체 직원의 휴가 현황이 관리자 뷰로 표시됨 (6 카테고리 그룹핑)'),
        (12, '팀별 휴가 현황 조회', 'MANAGER (manager@ctr.co.kr)', '로그인 완료', '',
         '팀 관리 > 팀 근태/휴가 클릭',
         '박준혁 팀의 휴가 현황(캘린더 또는 리스트)이 표시됨'),
        (13, '연차촉진 수동 통보 점검 (cron 미동작)', 'HR_ADMIN', '로그인 완료', '',
         '휴가 관리 > 사용촉진 로그 탭 확인 — 자동 발송 X, 인사담당자 수동 처리 필요',
         '사용촉진 통보 로그가 표시되지만 자동 발송은 미구현(매뉴얼 §9 #10 알려진 제약). 인사담당자가 수동으로 통보·이력 기록함을 확인'),
        (14, '지정연차 등록 + 자동 차감', 'HR_ADMIN', '로그인 완료',
         '지정일: 2026-08-12 (여름 집단휴가), 대상: 전체',
         '설정 > 휴가 > 지정연차 관리 > 신규 등록 > 날짜 + 대상 선택 > 저장',
         '지정연차가 등록되고 모든 직원의 가용 일수에서 1일이 자동 차감됨 (MyLeave KPI 반영)'),
        (15, '경조사 휴가 신청 (달력일 통산)', 'EMPLOYEE (employee-a@ctr.co.kr)', '로그인 완료',
         '본인결혼 5일, 시작일: 2026-06-15 (월)',
         '내 휴가 > 휴가 신청 > 카테고리 "경조사" > 본인결혼 선택 > 시작일 입력 > 제출',
         '본인결혼 5일이 달력일 + 통산(공휴일·주말 포함)으로 차감됨'),
        (16, '휴가 취소 — 사용 전', 'EMPLOYEE (employee-a@ctr.co.kr)', '승인된 미사용 휴가 존재', '',
         '내 휴가 > 승인된 미사용 휴가 선택 > 취소 요청',
         '휴가가 자동 취소되고 잔액이 즉시 복구됨'),
        (17, '휴가 취소 — 사용 중 (팀장 승인)', 'EMPLOYEE / MANAGER',
         '시작일 경과 + 종료일 미경과 휴가 존재', '',
         'EMPLOYEE: 취소 요청 → MANAGER: 결재함에서 승인',
         '사용일까지만 차감, 나머지는 복구. 팀장 승인 완료 후 반영'),
        (18, '연말 갱신 (이월/소멸)', 'HR_ADMIN', '로그인 완료', '',
         '인사관리 > 휴가/휴직 관리 > 연말 갱신 메뉴 확인',
         '연말 잔여 연차 정산 기능이 제공됨 (이월/소멸/수당전환 + 신규 부여)'),
    ]
    write_scenario_rows(ws, rows, mark_new={13, 14})
    add_dropdown(ws, f'I2:I{len(rows) + 1}', ['Pass', 'Fail', 'Block', 'N/A'])
    add_conditional_pass_fail(ws, f'I2:I{len(rows) + 1}')
    return len(rows)


# ============================================================
# 휴직 (LOA) — NEW SHEET
# ============================================================
def sheet_loa(wb):
    ws = wb.create_sheet('휴직')
    ws.sheet_properties.tabColor = SCENARIO_TAB
    apply_header(ws, SCENARIO_HEADERS, SCENARIO_WIDTHS)

    rows = [
        (1, '육아휴직 신청 (직원)', 'EMPLOYEE (employee-a@ctr.co.kr)', '로그인 완료',
         '휴직 유형: 육아휴직, 기간: 2026-08-01 ~ 2027-07-31',
         '나의 공간 > 휴직 신청 > 육아휴직 선택 > 기간 입력 > 사유 입력 > 제출',
         '휴직 신청이 접수되고 상태 "REQUESTED" → 결재 대기'),
        (2, 'HR 휴직 승인', 'HR_ADMIN (hr@ctr.co.kr)', 'Step 1 완료', '',
         '인사관리 > 휴가/휴직 관리 > 휴직 신청 목록 > Step 1 신청 선택 > 승인',
         '휴직이 승인되고 상태 "APPROVED" → "ACTIVE" 전환 (시작일 도래 시)'),
        (3, '휴직 상태 자동 전환', 'HR_ADMIN', 'Step 2 완료, 시작일 도래', '',
         '휴직 관리에서 해당 직원 상태 확인',
         '시작일에 자동으로 ACTIVE 전환되고 직원 상태 배지 "휴직 중" 표시'),
        (4, '출퇴근 대상 제외 확인', 'HR_ADMIN', 'Step 3 완료', '',
         '근태 관리에서 휴직 중 직원 검색',
         '해당 직원이 출퇴근 대상에서 자동 제외됨 (이상 항목 발생 안 함)'),
        (5, '급여 대상 자동 주입 확인', 'HR_ADMIN', 'Step 3 완료, 급여 실행 생성 시점', '',
         '급여 > 새 급여 실행 > 휴직 중 직원 포함 여부 확인',
         'ACTIVE LOA가 PayrollRun 생성 시점에 자동 주입되고 PayrollAdjustment 자동 생성됨'),
        (6, 'cross-month 일할계산', 'HR_ADMIN', '휴직 기간이 월 중간에 시작', '',
         '급여 실행에서 휴직 직원의 PayrollAdjustment 확인',
         '월별 일할계산이 자동 적용됨 (Phase 3 cross-month 자동화)'),
        (7, '복귀 알림 cron (D-7/D-3/D-1)', 'HR_ADMIN', '복귀 예정일 D-7', '',
         '대시보드 또는 알림 센터에서 복귀 알림 확인',
         'D-7/D-3/D-1 시점에 직원 본인 + HR에게 복귀 알림이 발송됨'),
        (8, '복직 신청 (직원 또는 HR 대리)', 'EMPLOYEE / HR_ADMIN', '복귀 예정일 도래', '',
         '나의 공간 > 휴직 신청 > 복직 신청 다이얼로그 또는 HR이 직원 상세에서 대리 신청',
         '복직 신청이 접수되고 결재 대기 상태'),
        (9, '복직 완료 처리', 'HR_ADMIN', 'Step 8 완료', '',
         '휴직 관리에서 복직 신청 선택 > 승인 / 복직 완료 처리',
         '상태 "COMPLETED" → 직원 상태 "ACTIVE" 복원 + 출퇴근/급여 대상 재포함'),
        (10, '복귀 후 PayrollAdjustment 역조정', 'HR_ADMIN', 'Step 9 완료, 다음 급여 실행', '',
         '다음 PayrollRun 생성 시 해당 직원의 Adjustment 확인',
         '복직 시점에 맞춰 일할계산 역조정이 자동 생성됨'),
        (11, '휴직 복귀 예정 직원 — 시드 데이터', 'HR_ADMIN', 'EDGE-006 시드 데이터 존재', '한복귀 (EDGE-006)',
         '인사관리 > 휴가/휴직 관리에서 한복귀 검색',
         '복귀 예정 알림이 표시되고 복귀 처리 버튼이 제공됨'),
        (12, '6-state 워크플로 확인', 'HR_ADMIN', '복수 휴직 건 존재', '',
         '휴직 관리 목록에서 상태 필터 확인',
         'REQUESTED / APPROVED / ACTIVE / RETURN_REQUESTED / COMPLETED / REJECTED 6개 상태 모두 표시 가능'),
    ]
    write_scenario_rows(ws, rows, mark_new=set(range(1, 13)))  # all new
    add_dropdown(ws, f'I2:I{len(rows) + 1}', ['Pass', 'Fail', 'Block', 'N/A'])
    add_conditional_pass_fail(ws, f'I2:I{len(rows) + 1}')
    return len(rows)


# ============================================================
# 급여
# ============================================================
def sheet_payroll(wb):
    ws = wb.create_sheet('급여')
    ws.sheet_properties.tabColor = SCENARIO_TAB
    apply_header(ws, SCENARIO_HEADERS, SCENARIO_WIDTHS)

    rows = [
        (1, '급여 대시보드 확인', 'HR_ADMIN (hr@ctr.co.kr)', '로그인 완료', '',
         '좌측 메뉴 > 급여 > 급여 대시보드 클릭',
         '급여 대시보드가 표시되고 최근 급여 실행 현황이 보임'),
        (2, '급여 실행 생성', 'HR_ADMIN', 'Step 1 완료', '급여월: 2026년 5월',
         '급여 대시보드 > 새 급여 실행 생성 > 2026년 5월 선택 > 생성',
         '5월 급여 실행이 "초안(DRAFT)" 상태로 생성됨'),
        (3, '출퇴근 마감 연동', 'HR_ADMIN', 'Step 2 완료', '',
         '급여 > 근태 마감 메뉴에서 5월 마감 상태 확인',
         '출퇴근 마감이 급여 실행과 연동되어 근무일수/초과근무가 반영됨'),
        (4, '급여 조정 입력 (수동 조정)', 'HR_ADMIN', 'Step 2 완료',
         '대상: 이민준, 조정: 야근수당 200,000원',
         '급여 > 수동 조정 메뉴 > 이민준 선택 > 야근수당 입력 > 저장',
         '조정 항목이 급여 실행에 반영됨'),
        (5, '급여 이상 감지', 'HR_ADMIN', 'Step 4 완료', '',
         '급여 > 이상 검토 메뉴 클릭',
         '전월 대비 큰 변동이 있는 직원이 6개 이상 감지 규칙으로 표시됨'),
        (6, '급여 시뮬레이션', 'HR_ADMIN', 'Step 2 완료', '',
         '급여 > 급여 시뮬레이션 메뉴에서 5월 급여 시뮬레이션 실행',
         '급여 시뮬레이션 결과가 표시되고 실제 실행 전 검증 가능'),
        (7, '급여 검토', 'HR_ADMIN', 'Step 6 완료', '',
         '급여 실행 > 검토 페이지에서 직원별 급여 항목 확인',
         '기본급, 식대, 4대보험 공제 등 항목별 금액이 정확히 표시됨'),
        (8, '급여 다단계 승인 (KR 2단계)', 'HR_ADMIN → MANAGER → CEO', 'Step 7 완료', '',
         '급여 실행 > 승인 요청 > 본부장 1차 승인 → 대표 최종 승인',
         '한국 법인 2단계 결재 완료. 9-state pipeline 진행 (DRAFT → ... → PAID)'),
        (9, '이체 파일 생성', 'HR_ADMIN', 'Step 8 완료', '',
         '급여 > 이체 내역 메뉴에서 은행 이체 파일 생성',
         '은행 이체용 CSV 파일이 생성되고 다운로드 가능'),
        (10, '급여명세서 발행', 'HR_ADMIN', 'Step 8 완료', '',
         '급여 실행 > 명세서 발행 클릭',
         '직원별 급여명세서가 생성되고 발행 상태가 "발행완료"로 변경됨'),
        (11, 'EMPLOYEE 급여명세서 확인', 'EMPLOYEE (employee-a@ctr.co.kr)', 'Step 10 완료', '',
         '나의 공간 > 급여명세서 클릭',
         '이민준의 5월 급여명세서가 표시되고 항목별 금액 확인 가능'),
        (12, '급여명세서 항목 확인', 'EMPLOYEE (employee-a@ctr.co.kr)', 'Step 11 완료', '',
         '급여명세서에서 세부 항목 확인',
         '기본급, 식대, 교통비, 4대보험(국민연금/건강보험/고용보험/산재보험) + 소득세 항목이 모두 표시됨'),
        (13, '글로벌 급여 관리 (해외 법인)', 'HR_ADMIN', '로그인 완료', '',
         '급여 > 글로벌 급여 메뉴 클릭',
         '해외 법인(CTR-CN 등)의 급여 관리 화면이 표시됨 (외부 처리 결과 업로드 방식)'),
        (14, '연말정산 (한국)', 'HR_ADMIN', '로그인 완료', '',
         '급여 > 연말정산 메뉴 클릭',
         '한국 법인 직원의 연말정산 관리 화면이 표시됨'),
        (15, '4 종 export 파일', 'HR_ADMIN', 'Step 10 완료', '',
         '급여 실행 상세에서 4종 export 확인',
         '비교(comparison) / 원장(ledger) / 분개(journal) / 은행 CSV 4종 모두 다운로드 가능'),
        (16, '무급 인턴 급여 제외', 'HR_ADMIN', 'EDGE-017 시드 데이터 존재', '무급실 (EDGE-017)',
         '급여 실행에서 무급실 확인',
         '무급 인턴이 급여 지급 대상에서 제외되거나 0원으로 처리됨'),
        (17, '연봉 밴드 초과자 확인', 'HR_ADMIN', 'EDGE-015 시드 데이터 존재', '고초과 (EDGE-015)',
         '보상 또는 급여 관리에서 밴드 초과자 확인',
         '고초과의 현재 연봉이 해당 직급 밴드 상한을 초과한다는 경고가 표시됨'),
        (18, '연봉 밴드 미달자 확인', 'HR_ADMIN', 'EDGE-016 시드 데이터 존재', '하미달 (EDGE-016)',
         '보상 또는 급여 관리에서 밴드 미달자 확인',
         '하미달의 현재 연봉이 해당 직급 밴드 하한에 미달한다는 경고가 표시됨'),
        (19, '빈번한 급여 변경 직원', 'HR_ADMIN', 'EDGE-018 시드 데이터 존재', '변봉급 (EDGE-018)',
         '급여 이상 검토에서 변봉급 확인',
         '잦은 급여 변경 이력이 이상 항목으로 감지됨'),
        (20, '비정기 조정 대기 확인', 'HR_ADMIN', 'EDGE-019 시드 데이터 존재', '승대기 (EDGE-019)',
         '보상 > 비정기 조정 목록에서 승대기 확인',
         '승인 대기 중인 비정기 조정 건이 표시됨'),
        (21, 'EMPLOYEE 타인 급여 접근 차단', 'EMPLOYEE (employee-a@ctr.co.kr)',
         '로그인 완료', 'URL: /payroll/me/[다른직원RunID]',
         '브라우저 주소창에 다른 직원의 급여명세서 URL 직접 입력',
         '403 접근 거부 또는 본인 급여만 조회 가능'),
        (22, '급여 실행 삭제/취소', 'HR_ADMIN', '초안 상태의 급여 실행 존재', '',
         '급여 대시보드에서 초안 상태 급여 실행 선택 > 삭제/취소',
         '초안 상태의 급여 실행이 삭제/취소됨'),
    ]
    write_scenario_rows(ws, rows)
    add_dropdown(ws, f'I2:I{len(rows) + 1}', ['Pass', 'Fail', 'Block', 'N/A'])
    add_conditional_pass_fail(ws, f'I2:I{len(rows) + 1}')
    return len(rows)


# ============================================================
# 결재·승인 (EXPANDED with new scenarios)
# ============================================================
def sheet_approval(wb):
    ws = wb.create_sheet('결재·승인')
    ws.sheet_properties.tabColor = SCENARIO_TAB
    apply_header(ws, SCENARIO_HEADERS, SCENARIO_WIDTHS)

    rows = [
        (1, 'EMPLOYEE 결재 요청 생성 (휴가)', 'EMPLOYEE (employee-a@ctr.co.kr)', '로그인 완료',
         '휴가 신청: 연차 1일 (2026-05-22)',
         '나의 공간 > 휴가 신청 > 제출',
         '결재 요청이 생성되고 매니저의 결재함에 전송됨'),
        (2, 'MANAGER 결재함 대기 건 확인', 'MANAGER (manager@ctr.co.kr)', 'Step 1 완료', '',
         '좌측 메뉴 > 결재 > 결재함 클릭',
         '이민준의 휴가 신청이 대기 건으로 표시됨'),
        (3, '결재 승인', 'MANAGER (manager@ctr.co.kr)', 'Step 2 완료', '',
         '결재함에서 이민준 휴가 신청 선택 > 승인',
         '결재가 승인되고 상태가 "완료"로 변경됨'),
        (4, '결재 반려 + 사유', 'MANAGER (manager@ctr.co.kr)', '새로운 결재 대기 건 존재',
         '반려 사유: 날짜 변경 요청',
         '결재함에서 대기 건 선택 > 반려 > 사유 입력 > 확인',
         '결재가 반려되고 신청자에게 반려 사유가 전달됨'),
        (5, '다단계 승인 흐름 (채용 결재)', 'HR_ADMIN', '채용요청서 시드 존재', '',
         '채용 > 채용요청서 > 새 요청서 작성 > 제출 → 다단계 승인 진행',
         '채용 1단계(본부장) 승인 후 2단계(전결)로 자동 전달 또는 finalize됨'),
        (6, 'dept_head 결재 라우팅', 'HR_ADMIN / dept_head', 'Department.head_employee_id 시드 존재',
         '52개 부서장 시드',
         '결재 라우팅에서 dept_head 단계 결재자 확인',
         'Department.head_employee_id 기반으로 부서장이 결재자로 자동 지정됨 (resolveApproverByRole)'),
        (7, 'multi-role canApprove (detail page)', 'multi-role 사용자',
         'HR_ADMIN base + 보조 MANAGER active EmployeeRole', '',
         '채용요청서 상세 페이지에서 결재 버튼 노출 확인',
         'multi-role employee의 detail page에서 결재 버튼이 false-deny되지 않고 정상 노출됨 (PR #18)'),
        (8, 'off-cycle approve trailing self-skip', 'HR_ADMIN → MANAGER',
         '[direct_manager → hr_admin] flow + HR_ADMIN 발의', '',
         'HR_ADMIN이 off-cycle 보상 발의 → step 2 self-skip APPROVED + step 1 PENDING → MANAGER가 step 1 결재',
         'step 1 결재 후 즉시 finalize되어 status=APPROVED, currentStep=totalSteps로 갱신 (PR #19 stuck fix)'),
        (9, 'secondary assignment 결재자', 'secondary manager', '직원에 보조 매니저 active', '',
         '보조 매니저로 결재 시도',
         '보조(secondary) assignment 결재자도 결재 가능. primary만 가능하지 않음 (Session 204)'),
        (10, 'pre-hire 로그인 차단', 'pre-hire 계정', 'effectiveDate > now', '',
         '입사일 전 신규 직원 계정으로 로그인 시도',
         'EMPLOYEE/[] 강등 적용되어 권한 없는 빈 세션 통과 차단 (Session 209)'),
        (11, '위임 설정', 'MANAGER (manager@ctr.co.kr)', '로그인 완료',
         '위임 대상: 김서연 (manager2), 기간: 2026-05-20~30',
         '팀 관리 > 위임 설정 > 새 위임 > 대리자/기간 지정 > 저장',
         '위임이 설정되고 해당 기간 동안 결재 건이 대리자에게 전달됨'),
        (12, '위임 결재 동작', 'MANAGER (manager2@ctr.co.kr)',
         'Step 11 완료, 위임 기간 중 결재 건 존재', '',
         '결재함에서 위임받은 결재 건 확인 > 승인',
         '위임 결재자(김서연)가 정상적으로 승인 처리 가능'),
        (13, 'HR_ADMIN 전체 결재 현황', 'HR_ADMIN (hr@ctr.co.kr)', '로그인 완료', '',
         '결재 > 결재함 또는 전체 현황 클릭',
         'HR_ADMIN이 전체 법인의 결재 현황(대기/완료/반려)을 조회할 수 있음'),
        (14, '결재 라인 설정 UI', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 시스템 > 결재 플로우 확인',
         '결재 유형별(휴가/출퇴근/급여 등) 결재 라인 등록 화면. ⚠ 휴가 모듈은 이 설정을 읽지 않음 (manuals/leave.md §5.2 알려진 제약)'),
        (15, '출퇴근 결재', 'MANAGER (manager@ctr.co.kr)', '출퇴근 수정 요청 결재 건 존재', '',
         '결재함 > 출퇴근 탭에서 수정 요청 확인 > 승인',
         '출퇴근 수정이 승인되고 기록에 반영됨'),
        (16, '급여 결재', 'HR_ADMIN', '급여 실행 승인 요청 건 존재', '',
         '결재함에서 급여 실행 승인 요청 확인 > 승인',
         '급여 실행이 최종 승인됨 (KR 2단계)'),
        (17, '비정기 보상 결재', 'HR_ADMIN', '비정기 보상 결재 건 존재', '',
         '결재함에서 비정기 보상 요청 확인 > 승인',
         '비정기 보상이 승인되고 해당 직원 급여에 반영됨 (Off-Cycle Workflow)'),
        (18, '결재 이력 조회', 'EMPLOYEE (employee-a@ctr.co.kr)', '이전 결재 건 존재', '',
         '결재 > 내 신청 내역 확인',
         '본인이 신청한 결재 건의 이력(상태, 승인일, 결재자)이 표시됨'),
        (19, '결재 철회', 'EMPLOYEE (employee-a@ctr.co.kr)', '승인 대기 중인 결재 건 존재', '',
         '내 신청 내역에서 대기 건 선택 > 철회',
         '결재 요청이 철회되고 상태가 "철회"로 변경됨'),
        (20, '매니저 없는 직원 결재 라인', 'HR_ADMIN', 'EDGE-011 시드 데이터 존재', '윤대표 (EDGE-011)',
         '윤대표(상위 보고자 없음)의 결재 라인 확인',
         '매니저 없는 직원의 결재가 스킵되거나 HR로 직접 전달됨'),
        (21, '동시 결재 race protection', 'MANAGER × 2', '동일 결재 건 더블 클릭',
         '같은 사용자가 동시에 결재 버튼 2회 클릭 또는 mixed approve/reject',
         '결재함에서 같은 건 빠른 연속 클릭',
         'updateMany + status="PENDING" race protection으로 첫 클릭만 처리됨 (Session 211 PR #20)'),
        (22, '일괄 결재', 'MANAGER (manager@ctr.co.kr)', '복수 결재 대기 건 존재', '',
         '결재함에서 복수 건 선택 > 일괄 승인',
         '선택한 건들이 일괄로 승인 처리됨'),
    ]
    write_scenario_rows(ws, rows, mark_new={6, 7, 8, 9, 10, 21})
    add_dropdown(ws, f'I2:I{len(rows) + 1}', ['Pass', 'Fail', 'Block', 'N/A'])
    add_conditional_pass_fail(ws, f'I2:I{len(rows) + 1}')
    return len(rows)


# ============================================================
# 마스터데이터·설정
# ============================================================
def sheet_settings(wb):
    ws = wb.create_sheet('마스터데이터·설정')
    ws.sheet_properties.tabColor = SCENARIO_TAB
    apply_header(ws, SCENARIO_HEADERS, SCENARIO_WIDTHS)

    rows = [
        (1, '법인 기본정보 확인', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 조직 > 법인 기본정보 확인', '법인 코드, 사업자번호, 주소 등 기본정보가 정확히 표시됨'),
        (2, '법인 정보 수정', 'HR_ADMIN', 'Step 1 완료', '변경: 대표번호 수정',
         '법인 정보에서 대표번호 수정 > 저장', '변경사항이 저장되고 반영됨'),
        (3, '부서 구조 조회', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 조직 > 부서 관리 클릭', '부서 트리 구조가 표시되고 상위-하위 관계가 정확함'),
        (4, '부서 추가', 'HR_ADMIN', 'Step 3 완료', '새 부서: AI연구팀 (상위: 개발부)',
         '부서 관리 > 부서 추가 > 부서명, 상위부서 입력 > 저장',
         '새 부서가 생성되고 트리에 올바른 위치에 표시됨'),
        (5, '부서장 지정 (Department.head_employee_id)', 'HR_ADMIN', 'Step 3 완료',
         '대상 부서: QA-TEAM-A, 부서장: 박준혁',
         '부서 관리 > 해당 부서 선택 > 부서장 지정 > 박준혁 선택',
         'Department.head_employee_id가 박준혁으로 설정되고 dept_head 결재 라우팅에 반영됨'),
        (6, '직위 관리', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 조직 > 직위 관리 확인',
         '직위 목록(대표/본부장/팀장/책임/선임/사원 등)이 표시됨'),
        (7, '직급 체계 확인 (Session 167 전환)', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 조직 > 직급 관리에서 L/E/S 체계 확인',
         '국내 직급 E1/S1/L2/L1 (4단계) + 해외 직급 L1~L5/S1/E1 (7단계)이 정확히 표시됨'),
        (8, '직종/직무 관리', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 조직 > 직종/직무 관리 확인', '직종(기술, 경영 등)과 직무 목록이 표시됨'),
        (9, '발령 규칙 확인', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 조직 > 발령 규칙 확인',
         '발령 유형별 규칙(승진 요건, 이동 제한 등)이 설정되어 있음'),
        (10, '수습 기간 설정', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 조직 > 수습 기간 설정 확인', '수습 기간(3개월/6개월 등)이 설정되어 있음'),
        (11, '커스텀 필드 관리', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 조직 > 커스텀 필드 관리 확인',
         '사용자 정의 필드 목록이 표시되고 추가/수정 가능'),
        (12, '코드 관리', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 조직 > 코드 관리 확인',
         '시스템 코드(고용형태, 발령유형 등) 목록이 관리 가능'),
        (13, '근무지 관리', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 조직 > 근무지 관리 확인',
         '근무지 목록(본사, 지사 등)이 표시되고 위치 정보 확인 가능'),
        (14, '근무 스케줄 확인', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 출퇴근 > 근무 스케줄 확인',
         '기본 근무 스케줄(출근 09:00, 퇴근 18:00 등)이 설정되어 있음'),
        (15, '근무 스케줄 수정', 'HR_ADMIN', 'Step 14 완료',
         '변경: 유연근무제 점심시간 12:00~13:00',
         '근무 스케줄 수정 > 저장', '변경된 스케줄이 저장되고 반영됨'),
        (16, '주 52시간 한도 설정', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 출퇴근 > 근무한도에서 52시간 설정 확인',
         '주 52시간 한도가 설정되어 있고 경고 임계값이 확인됨'),
        (17, '교대근무 패턴', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 출퇴근 > 교대근무 패턴 확인',
         '교대근무 패턴(주간/야간/3교대 등)이 설정되어 있음'),
        (18, '휴가 유형 확인', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 휴가 > 휴가 유형 관리 확인',
         '연차, 병가, 경조사, 출산휴가 등 휴가 유형이 6 카테고리로 그룹핑되어 표시됨'),
        (19, '휴가 부여 규칙', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 휴가 > 부여 규칙 확인', '근속연수별 연차 부여 규칙이 설정되어 있음'),
        (20, '지정연차 관리', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 휴가 > 지정연차 관리 확인',
         '회사 지정 연차일 등록/관리 기능이 제공됨 (DesignatedLeaveDay)'),
        (21, '법정 공휴일 캘린더', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 휴가 > 공휴일 캘린더 확인',
         '한국 법정 공휴일이 등록되어 있고 연도별 관리 가능 (해외 5개 법인은 매뉴얼 §9 #2 알려진 제약)'),
        (22, '초과근무 설정', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 출퇴근 > 초과근무 설정 확인',
         '초과근무 승인 방식, 수당 계산 규칙이 설정되어 있음'),
        (23, '휴직 유형 관리', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 휴가 > 휴직 유형 관리 확인',
         '육아휴직, 병간호휴직, 개인사유 등 휴직 유형이 관리 가능 (LeaveOfAbsenceType)'),
        (24, '급여 항목 관리', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 급여 > 급여 항목 관리 확인',
         '기본급, 식대, 교통비 등 급여 항목이 목록으로 표시됨'),
        (25, '공제 항목 관리', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 급여 > 공제 항목 관리 확인',
         '국민연금, 건강보험, 고용보험, 산재보험, 소득세 등 공제 항목이 표시됨'),
        (26, '비과세 한도 확인', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 급여 > 비과세 한도 확인',
         '식대, 교통비 등 비과세 한도가 설정되어 있음'),
        (27, '연봉 밴드 확인', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 급여 > 연봉 밴드 확인',
         '직급별 연봉 밴드(최소~최대) 범위가 설정되어 있음'),
        (28, '인상률 매트릭스', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 급여 > 인상률 매트릭스 확인',
         '성과등급별 인상률 매트릭스가 설정되어 있음'),
        (29, '성과급 규칙', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 급여 > 성과급 규칙 확인',
         '성과등급별 성과급 비율/금액 규칙이 설정되어 있음'),
        (30, '급여일 설정', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 급여 > 급여일 설정 확인', '매월 급여일(25일 등)이 설정되어 있음'),
        (31, '통화/환율 관리', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 급여 > 통화/환율 관리 확인',
         '해외 법인 통화(CNY, USD 등)와 환율이 관리 가능'),
        (32, '성과 평가 주기 설정', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 성과 > 평가 주기 설정 확인', '연간/반기/분기 등 평가 주기가 설정되어 있음'),
        (33, '평가 항목 설정', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 성과 > 평가 항목/가중치 확인',
         '역량평가, 업적평가 등 항목과 가중치가 설정되어 있음'),
        (34, '채용 파이프라인 설정', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 채용 > 채용 파이프라인 설정 확인',
         '10-stage 파이프라인(서류→면접→오퍼→OFFER_ACCEPTED/DECLINED 등)이 설정되어 있음'),
        (35, '시스템 결재 플로우 설정', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 시스템 > 결재 플로우 확인',
         '결재 유형별 결재 라인과 규칙이 설정되어 있음 (휴가 모듈 미연동 명시)'),
        (36, '알림 설정', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 시스템 > 알림 설정 확인',
         '이메일/인앱/Teams/푸시 알림 유형별 설정이 관리 가능'),
        (37, '컴플라이언스 허브', 'HR_ADMIN', '로그인 완료', '',
         '설정 > 컴플라이언스 클릭',
         'GDPR / DPIA / 데이터 보존 / PII 감사 / 국가별 노동법 준수 화면 진입 가능'),
    ]
    write_scenario_rows(ws, rows, mark_new={5, 7, 20})
    add_dropdown(ws, f'I2:I{len(rows) + 1}', ['Pass', 'Fail', 'Block', 'N/A'])
    add_conditional_pass_fail(ws, f'I2:I{len(rows) + 1}')
    return len(rows)


# ============================================================
# Continued in next call — split scripts for size
# ============================================================
if __name__ == '__main__':
    # Build main workbook
    from build_testbook_v2_part2 import attach_part2

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    sheet_usage(wb)
    n_employee = sheet_employee(wb)
    n_attendance = sheet_attendance(wb)
    n_leave = sheet_leave(wb)
    n_loa = sheet_loa(wb)
    n_payroll = sheet_payroll(wb)
    n_approval = sheet_approval(wb)
    n_settings = sheet_settings(wb)

    counts = {
        '인사관리': n_employee,
        '출퇴근': n_attendance,
        '휴가관리': n_leave,
        '휴직': n_loa,
        '급여': n_payroll,
        '결재·승인': n_approval,
        '마스터데이터·설정': n_settings,
    }

    attach_part2(wb, counts)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    size_kb = OUT_PATH.stat().st_size / 1024
    print(f'✓ Wrote {OUT_PATH.relative_to(REPO_ROOT)} ({size_kb:.1f} KB)')
    print(f'  Sheets: {len(wb.sheetnames)}')
    for sn in wb.sheetnames:
        print(f'    - {sn}')
