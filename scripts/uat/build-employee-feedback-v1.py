#!/usr/bin/env python3
"""
CTR HR Hub — 사무직 친화 UAT 피드백 폼 v1 generator

Run: python3 scripts/uat/build-employee-feedback-v1.py
Output: docs/uat/UAT_피드백_사무직_v1.xlsx

Design goals:
- 30초 안에 한 건 입력 가능 (컬럼 7개)
- 한글 친화 표현 (Critical → "🔴 일이 안 돼요")
- 버그 신고 vs 개선 제안 시트 분리 (분류 부담 제거)
- 이슈ID·담당자·결함경과일 등 분류 메타 제거 (UAT 관리자가 v2로 매핑)
- 시트 색상 분리 (빨강 = 문제, 노랑 = 제안, 회색 = 안내, 보라 = 관리자 집계)

기존 빌더 패턴 재활용: scripts/uat/build-feedback-v2.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[2]
OUT_PATH = REPO_ROOT / 'docs/uat/UAT_피드백_사무직_v1.xlsx'

# ============================================================
# Styles
# ============================================================
FONT_NAME = 'Malgun Gothic'
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

# Headers per sheet color theme
HEADER_FONT = Font(name=FONT_NAME, size=12, bold=True, color='FFFFFF')

# Sheet tab colors
TAB_GUIDE = '808080'       # gray
TAB_PROBLEM = 'C00000'     # red
TAB_SUGGESTION = 'FFC000'  # yellow
TAB_ADMIN = '7030A0'       # purple

# Header fills per sheet
FILL_PROBLEM = PatternFill('solid', fgColor='C00000')
FILL_SUGGESTION = PatternFill('solid', fgColor='BF8F00')  # darker yellow for contrast
FILL_ADMIN = PatternFill('solid', fgColor='7030A0')
FILL_GUIDE = PatternFill('solid', fgColor='595959')

# Body styles
TITLE_FONT = Font(name=FONT_NAME, size=18, bold=True, color='1F4E78')
SECTION_FONT = Font(name=FONT_NAME, size=14, bold=True, color='1F4E78')
BODY_FONT = Font(name=FONT_NAME, size=12)
BIG_BODY_FONT = Font(name=FONT_NAME, size=13)
EXAMPLE_GOOD_FILL = PatternFill('solid', fgColor='E2EFDA')  # soft green
EXAMPLE_BAD_FILL = PatternFill('solid', fgColor='FCE4D6')   # soft red
CALLOUT_FILL = PatternFill('solid', fgColor='FFF2CC')       # soft yellow

ALIGN_TOP_WRAP = Alignment(vertical='top', wrap_text=True)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_CENTER_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ============================================================
# Dropdown values — 한글 친화
# ============================================================
SCREENS = [
    '홈 / 알림',
    '나의 공간 (내 정보)',
    '휴가 신청·조회',
    '휴직 신청·조회',
    '출퇴근',
    '결재함 (받은 결재)',
    '내 결재 신청 내역',
    '급여명세서',
    '연말 정산',
    '내 목표·평가',
    '분기 리뷰',
    '인정·칭찬',
    '리포트·통계',
    '직원 관리 (HR)',
    '조직도·부서 관리 (HR)',
    '근태 관리 (HR)',
    '휴가 관리 (HR)',
    '급여 관리 (HR)',
    '결재 흐름 설정 (HR)',
    '채용 관리 (HR)',
    '평가 관리 (HR)',
    '보상 관리 (HR)',
    '온보딩·오프보딩 (HR)',
    '마스터데이터·설정 (HR)',
    '컴플라이언스 (HR)',
    '기타',
]

# 3-level severity (problem)
PROBLEM_LEVELS = [
    '🔴 일이 안 돼요 (업무 불가능)',
    '🟡 불편하지만 우회 가능',
    '🟢 사소해요 (별 문제 없음)',
]

# 3-level priority (suggestion)
SUGGESTION_PRIORITY = [
    '⭐⭐⭐ 매일 도움이 될 거예요',
    '⭐⭐ 가끔 도움이 될 거예요',
    '⭐ 있으면 좋겠어요',
]


# ============================================================
# Helpers
# ============================================================
def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_header_row(ws, headers, fill, row=1):
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = fill
        c.alignment = ALIGN_CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 36
    ws.freeze_panes = f'A{row + 1}'


def write_empty_data_rows(ws, n_rows, n_cols, start_row=2, height=80):
    """Pre-format N empty rows for HR testers to fill."""
    for r in range(start_row, start_row + n_rows):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BIG_BODY_FONT
            cell.alignment = ALIGN_TOP_WRAP
            cell.border = BORDER
        ws.row_dimensions[r].height = height


# ============================================================
# Sheet 1: 시작하기 (안내)
# ============================================================
def sheet_guide(wb):
    ws = wb.create_sheet('🚀 시작하기')
    ws.sheet_properties.tabColor = TAB_GUIDE

    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 100

    # Title
    ws['B1'] = '🚀 시작하기 — UAT 피드백 작성 가이드'
    ws['B1'].font = TITLE_FONT
    ws.merge_cells('B1:B1')
    ws.row_dimensions[1].height = 36

    # Intro
    intro_rows = [
        '',
        '안녕하세요! CTR HR Hub UAT 테스트에 참여해주셔서 감사합니다.',
        '테스트 중 발견한 문제나 아이디어가 있으면 본 파일에 편하게 적어주세요.',
        '',
        '📌 작성 원칙 한 줄: "다음 사람이 그대로 따라해서 재현할 수 있게 적어주세요."',
        '',
    ]
    r = 2
    for line in intro_rows:
        if line:
            ws.cell(row=r, column=2, value=line).font = BIG_BODY_FONT
        r += 1

    # Section: How to use
    ws.cell(row=r, column=2, value='1️⃣ 어떻게 사용하나요?').font = SECTION_FONT
    r += 1
    how_rows = [
        '• 화면 아래 탭(시트)에서 어떤 종류인지 고르고 → 한 줄(한 행)에 7개 칸을 채워주세요.',
        '   🔴 이건 잘 안돼요  — 문제·버그·오류·이상한 동작 신고',
        '   💡 이게 있으면 좋겠어요  — 새 기능·개선·편의성 제안',
        '• 한 건당 30초 안에 적고 다음 테스트로 돌아가세요. 너무 자세히 안 적어도 OK.',
        '• 스크린샷은 선택사항이지만, 가능하면 첨부해주시면 개발팀이 빨리 이해해요. (PrintScreen → 셀에 붙여넣기)',
        '• 같은 문제 여러 번 보였으면 새 줄 말고 비고에 횟수만 적어주세요.',
    ]
    for line in how_rows:
        ws.cell(row=r, column=2, value=line).font = BIG_BODY_FONT
        ws.cell(row=r, column=2).alignment = ALIGN_TOP_WRAP
        ws.row_dimensions[r].height = 24
        r += 1

    r += 1
    # Section: Severity meaning
    ws.cell(row=r, column=2, value='2️⃣ 색깔·별점은 어떻게 정해요?').font = SECTION_FONT
    r += 1
    severity_rows = [
        ('🔴 일이 안 돼요', '버튼 눌러도 작동 안 함, 화면이 안 뜸, 잘못된 정보가 저장됨 → 업무 자체 불가능'),
        ('🟡 불편하지만 우회 가능', '느리거나 한 번에 안 되지만 다른 방법으로는 됨 → 짜증나지만 일은 됨'),
        ('🟢 사소해요', '화면 깨짐, 글자 색깔 이상, 단순 오타 → 거슬리지만 일에는 지장 X'),
        ('⭐⭐⭐ 매일 도움', '있으면 매일 쓸 정도로 자주 필요한 기능'),
        ('⭐⭐ 가끔 도움', '한 달에 몇 번 쓸 기능'),
        ('⭐ 있으면 좋겠어요', '드물게 필요하지만 있으면 좋겠다 싶은 기능'),
    ]
    for level, desc in severity_rows:
        c1 = ws.cell(row=r, column=2, value=f'{level}  —  {desc}')
        c1.font = BIG_BODY_FONT
        c1.alignment = ALIGN_TOP_WRAP
        ws.row_dimensions[r].height = 24
        r += 1

    r += 1
    # Section: Good examples
    ws.cell(row=r, column=2, value='3️⃣ 좋은 예시 (이렇게 적어주세요)').font = SECTION_FONT
    r += 1

    good_examples = [
        {
            '제목': '✅ 좋은 예시 1 — 반차 잔액 차감 오류',
            '화면': '휴가 신청·조회',
            '무슨 일': '2026-05-25 오전 09:00~13:00 반차를 신청하고 매니저 승인까지 받았는데, 잔여 연차가 1일 통째로 차감되었습니다. (반차니까 0.5일만 차감되어야 할 것 같아요)',
            '어떻게': '반차는 0.5일만 차감되어야 합니다. 잔여 연차 표시도 14.5일로 보여줘야 할 것 같아요.',
            '얼마나': '🔴 일이 안 돼요',
        },
        {
            '제목': '✅ 좋은 예시 2 — 결재함 알림 안 옴',
            '화면': '결재함 (받은 결재)',
            '무슨 일': '팀원이 어제 휴가 신청을 했는데 제 결재함에 빨간 점(알림)이 안 떴어요. 직접 결재함을 들어가서야 신청 건이 있는 걸 봤어요.',
            '어떻게': '팀원이 결재 요청하면 사이드바 결재함 메뉴에 빨간 점이 떠야 한다고 생각해요.',
            '얼마나': '🟡 불편하지만 우회 가능',
        },
        {
            '제목': '💡 좋은 예시 3 — 휴가 신청 화면 개선 아이디어',
            '화면': '휴가 신청·조회',
            '무슨 일/어떤 점': '휴가 신청한 다음에 누가 결재했는지, 지금 어느 단계인지 한눈에 안 보여요. 매번 상세 페이지를 클릭해서 들어가야 알 수 있어요.',
            '어떻게': '신청 목록에 결재자 이름이랑 진행 단계(승인 대기 / 1차 승인 완료 / 최종 승인 등)를 같이 표시해주세요.',
            '얼마나': '⭐⭐⭐ 매일 도움이 될 거예요',
        },
    ]

    for ex in good_examples:
        for k, v in ex.items():
            cell = ws.cell(row=r, column=2, value=f'  {k}: {v}')
            cell.font = BIG_BODY_FONT
            cell.fill = EXAMPLE_GOOD_FILL
            cell.alignment = ALIGN_TOP_WRAP
            ws.row_dimensions[r].height = 28
            r += 1
        r += 1

    # Section: Bad examples
    ws.cell(row=r, column=2, value='4️⃣ 나쁜 예시 (이러면 개발팀이 못 고쳐요)').font = SECTION_FONT
    r += 1

    bad_examples = [
        {
            '제목': '❌ 나쁜 예시 1 — 너무 추상적',
            '화면': '(비어있음)',
            '무슨 일': '휴가가 이상해요',
            '어떻게': '고쳐주세요',
            '왜 나쁜가': '→ 어느 화면에서, 어떤 휴가가, 어떻게 이상한지 모름. 개발팀이 재현 못해서 fix 불가능.',
        },
        {
            '제목': '❌ 나쁜 예시 2 — 감정만',
            '화면': '결재함 (받은 결재)',
            '무슨 일': '왜 이렇게 느려요?',
            '어떻게': '빨리 해주세요',
            '왜 나쁜가': '→ "느리다"는 게 어느 정도(5초? 30초?), 어느 액션에서, 어떤 데이터에서 그런지 모름.',
        },
    ]

    for ex in bad_examples:
        for k, v in ex.items():
            cell = ws.cell(row=r, column=2, value=f'  {k}: {v}')
            cell.font = BIG_BODY_FONT
            cell.fill = EXAMPLE_BAD_FILL
            cell.alignment = ALIGN_TOP_WRAP
            ws.row_dimensions[r].height = 28
            r += 1
        r += 1

    # Section: FAQ
    ws.cell(row=r, column=2, value='5️⃣ 자주 묻는 질문').font = SECTION_FONT
    r += 1
    faq = [
        ('Q. 한 건당 얼마나 적어야 해요?', 'A. 3~5줄이면 충분해요. 길게 적을 필요 없어요. 다만 화면 이름과 "무슨 일이 있었나요"는 꼭 채워주세요.'),
        ('Q. 스크린샷은 꼭 첨부해야 하나요?', 'A. 선택사항입니다. 화면 깨짐·오류 메시지처럼 그림이 도움 되는 건만 첨부해주세요. PrintScreen 후 셀에 Ctrl+V로 붙여넣기.'),
        ('Q. 한 건인지 두 건인지 모르겠어요.', 'A. 일단 따로 적어주세요. 관리자가 정리할 때 합치면 돼요.'),
        ('Q. 다른 사람이 이미 신고한 것 같아요.', 'A. 그래도 적어주세요. 같은 문제를 여러 사람이 신고하면 우선순위가 올라가요.'),
        ('Q. 비밀번호 같은 민감 정보가 들어 있어요.', 'A. 스크린샷에서 민감 정보는 가려주세요(빨간 박스로 칠하기 등). 텍스트에도 적지 마세요.'),
    ]
    for q, a in faq:
        ws.cell(row=r, column=2, value=q).font = Font(name=FONT_NAME, size=12, bold=True)
        ws.row_dimensions[r].height = 22
        r += 1
        ws.cell(row=r, column=2, value=a).font = BIG_BODY_FONT
        ws.cell(row=r, column=2).alignment = ALIGN_TOP_WRAP
        ws.row_dimensions[r].height = 24
        r += 1
        r += 1

    # Closing callout
    callout = ws.cell(row=r, column=2,
                      value='💌 도움이 필요하면 언제든 UAT 관리자에게 직접 문의하세요. 완벽하게 안 적어도 괜찮습니다 — '
                            '"있는 그대로" 적어주시는 게 가장 도움이 됩니다.')
    callout.font = BIG_BODY_FONT
    callout.fill = CALLOUT_FILL
    callout.alignment = ALIGN_TOP_WRAP
    ws.row_dimensions[r].height = 40

    # Set view zoom level
    ws.sheet_view.zoomScale = 110


# ============================================================
# Sheet 2: 🔴 이건 잘 안돼요 (문제 신고)
# ============================================================
def sheet_problem(wb):
    ws = wb.create_sheet('🔴 이건 잘 안돼요')
    ws.sheet_properties.tabColor = TAB_PROBLEM

    headers = [
        '이름',
        '발견 날짜',
        '어느 화면에서?',
        '무슨 일이 있었나요?',
        '어떻게 됐으면 좋겠나요?',
        '얼마나 불편한가요?',
        '스크린샷 (선택)',
    ]
    widths = [12, 14, 24, 55, 45, 28, 25]
    set_col_widths(ws, widths)
    write_header_row(ws, headers, FILL_PROBLEM)

    # 50 empty rows
    write_empty_data_rows(ws, 50, len(headers), start_row=2, height=80)

    # Dropdown — 어느 화면 (col C)
    dv_screen = DataValidation(type='list', formula1=f'"{",".join(SCREENS)}"', allow_blank=True)
    dv_screen.add('C2:C51')
    ws.add_data_validation(dv_screen)

    # Dropdown — 얼마나 불편 (col F)
    dv_level = DataValidation(type='list', formula1=f'"{",".join(PROBLEM_LEVELS)}"', allow_blank=True)
    dv_level.add('F2:F51')
    ws.add_data_validation(dv_level)

    # Conditional formatting on severity
    ws.conditional_formatting.add(
        'F2:F51',
        CellIsRule(operator='equal', formula=[f'"{PROBLEM_LEVELS[0]}"'],
                   fill=PatternFill('solid', fgColor='FFC7CE'),
                   font=Font(name=FONT_NAME, size=12, bold=True, color='9C0006')))
    ws.conditional_formatting.add(
        'F2:F51',
        CellIsRule(operator='equal', formula=[f'"{PROBLEM_LEVELS[1]}"'],
                   fill=PatternFill('solid', fgColor='FFEB9C'),
                   font=Font(name=FONT_NAME, size=12, color='9C5700')))
    ws.conditional_formatting.add(
        'F2:F51',
        CellIsRule(operator='equal', formula=[f'"{PROBLEM_LEVELS[2]}"'],
                   fill=PatternFill('solid', fgColor='C6EFCE'),
                   font=Font(name=FONT_NAME, size=12, color='006100')))

    # Auto-date hint row (top sticky help)
    # Add a comment-like row at row 1 — actually we used row 1 as header. Add a help row at top? Nope.
    # Instead, set a sample row (row 2) lightly styled as placeholder hint
    # Skip — guide sheet has examples.

    ws.sheet_view.zoomScale = 110


# ============================================================
# Sheet 3: 💡 이게 있으면 좋겠어요 (개선 제안)
# ============================================================
def sheet_suggestion(wb):
    ws = wb.create_sheet('💡 이게 있으면 좋겠어요')
    ws.sheet_properties.tabColor = TAB_SUGGESTION

    headers = [
        '이름',
        '작성 날짜',
        '어느 화면?',
        '어떤 점이 아쉬워요?',
        '어떻게 해주세요?',
        '얼마나 도움이 될까요?',
        '스크린샷 (선택)',
    ]
    widths = [12, 14, 24, 55, 45, 28, 25]
    set_col_widths(ws, widths)
    write_header_row(ws, headers, FILL_SUGGESTION)

    write_empty_data_rows(ws, 50, len(headers), start_row=2, height=80)

    dv_screen = DataValidation(type='list', formula1=f'"{",".join(SCREENS)}"', allow_blank=True)
    dv_screen.add('C2:C51')
    ws.add_data_validation(dv_screen)

    dv_pri = DataValidation(type='list', formula1=f'"{",".join(SUGGESTION_PRIORITY)}"', allow_blank=True)
    dv_pri.add('F2:F51')
    ws.add_data_validation(dv_pri)

    # Conditional formatting on priority — soft, not warning
    ws.conditional_formatting.add(
        'F2:F51',
        CellIsRule(operator='equal', formula=[f'"{SUGGESTION_PRIORITY[0]}"'],
                   fill=PatternFill('solid', fgColor='D4E9D4'),
                   font=Font(name=FONT_NAME, size=12, bold=True, color='1F4E78')))
    ws.conditional_formatting.add(
        'F2:F51',
        CellIsRule(operator='equal', formula=[f'"{SUGGESTION_PRIORITY[1]}"'],
                   fill=PatternFill('solid', fgColor='E5E5E5'),
                   font=Font(name=FONT_NAME, size=12)))
    ws.conditional_formatting.add(
        'F2:F51',
        CellIsRule(operator='equal', formula=[f'"{SUGGESTION_PRIORITY[2]}"'],
                   fill=PatternFill('solid', fgColor='F5F5F5'),
                   font=Font(name=FONT_NAME, size=12, color='595959')))

    ws.sheet_view.zoomScale = 110


# ============================================================
# Sheet 4: 📊 관리자용 집계 (자동 COUNTIFS)
# ============================================================
def sheet_admin(wb):
    ws = wb.create_sheet('📊 관리자용 집계')
    ws.sheet_properties.tabColor = TAB_ADMIN

    ws.column_dimensions['A'].width = 36
    for c in range(2, 8):
        ws.column_dimensions[get_column_letter(c)].width = 16

    ws['A1'] = '📊 관리자용 집계 (UAT 관리자만 보세요)'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:G1')
    ws.row_dimensions[1].height = 28

    note = ws.cell(row=2, column=1,
                   value='⚠️ HR 테스터는 이 시트를 채울 필요가 없습니다. 다른 시트에 입력하시면 여기서 자동 집계됩니다.')
    note.font = BODY_FONT
    note.fill = CALLOUT_FILL
    note.alignment = ALIGN_TOP_WRAP
    ws.merge_cells('A2:G2')
    ws.row_dimensions[2].height = 26

    # Section 1: 문제 신고 — 화면별
    r = 4
    ws.cell(row=r, column=1, value='1️⃣ 문제 신고 — 화면별 건수').font = SECTION_FONT
    r += 1
    sheet_problem_name = "'🔴 이건 잘 안돼요'"
    sheet_suggest_name = "'💡 이게 있으면 좋겠어요'"

    headers = ['화면', '🔴 일이 안 돼요', '🟡 불편하지만 우회', '🟢 사소', '합계']
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = FILL_ADMIN
        c.alignment = ALIGN_CENTER
        c.border = BORDER
    ws.row_dimensions[r].height = 24
    r += 1

    for screen in SCREENS:
        ws.cell(row=r, column=1, value=screen).font = BODY_FONT
        # Critical
        ws.cell(row=r, column=2, value=f'=COUNTIFS({sheet_problem_name}!C2:C51,A{r},{sheet_problem_name}!F2:F51,"{PROBLEM_LEVELS[0]}")').font = BODY_FONT
        # Major
        ws.cell(row=r, column=3, value=f'=COUNTIFS({sheet_problem_name}!C2:C51,A{r},{sheet_problem_name}!F2:F51,"{PROBLEM_LEVELS[1]}")').font = BODY_FONT
        # Minor
        ws.cell(row=r, column=4, value=f'=COUNTIFS({sheet_problem_name}!C2:C51,A{r},{sheet_problem_name}!F2:F51,"{PROBLEM_LEVELS[2]}")').font = BODY_FONT
        # Total
        ws.cell(row=r, column=5, value=f'=SUM(B{r}:D{r})').font = BODY_FONT
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = ALIGN_TOP_WRAP
        r += 1

    # Sum row
    total_row = r
    ws.cell(row=total_row, column=1, value='합계').font = HEADER_FONT
    ws.cell(row=total_row, column=1).fill = FILL_ADMIN
    for c in range(2, 6):
        col_letter = get_column_letter(c)
        cell = ws.cell(row=total_row, column=c,
                       value=f'=SUM({col_letter}{total_row - len(SCREENS)}:{col_letter}{total_row - 1})')
        cell.font = HEADER_FONT
        cell.fill = FILL_ADMIN
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER
    ws.cell(row=total_row, column=1).border = BORDER
    ws.cell(row=total_row, column=1).alignment = ALIGN_CENTER

    # Section 2: 개선 제안 — 화면별
    r = total_row + 3
    ws.cell(row=r, column=1, value='2️⃣ 개선 제안 — 화면별 건수').font = SECTION_FONT
    r += 1

    headers = ['화면', '⭐⭐⭐ 매일', '⭐⭐ 가끔', '⭐ 있으면 좋음', '합계']
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = FILL_ADMIN
        c.alignment = ALIGN_CENTER
        c.border = BORDER
    ws.row_dimensions[r].height = 24
    r += 1

    for screen in SCREENS:
        ws.cell(row=r, column=1, value=screen).font = BODY_FONT
        ws.cell(row=r, column=2, value=f'=COUNTIFS({sheet_suggest_name}!C2:C51,A{r},{sheet_suggest_name}!F2:F51,"{SUGGESTION_PRIORITY[0]}")').font = BODY_FONT
        ws.cell(row=r, column=3, value=f'=COUNTIFS({sheet_suggest_name}!C2:C51,A{r},{sheet_suggest_name}!F2:F51,"{SUGGESTION_PRIORITY[1]}")').font = BODY_FONT
        ws.cell(row=r, column=4, value=f'=COUNTIFS({sheet_suggest_name}!C2:C51,A{r},{sheet_suggest_name}!F2:F51,"{SUGGESTION_PRIORITY[2]}")').font = BODY_FONT
        ws.cell(row=r, column=5, value=f'=SUM(B{r}:D{r})').font = BODY_FONT
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = ALIGN_TOP_WRAP
        r += 1

    total_row2 = r
    ws.cell(row=total_row2, column=1, value='합계').font = HEADER_FONT
    ws.cell(row=total_row2, column=1).fill = FILL_ADMIN
    for c in range(2, 6):
        col_letter = get_column_letter(c)
        cell = ws.cell(row=total_row2, column=c,
                       value=f'=SUM({col_letter}{total_row2 - len(SCREENS)}:{col_letter}{total_row2 - 1})')
        cell.font = HEADER_FONT
        cell.fill = FILL_ADMIN
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER
    ws.cell(row=total_row2, column=1).border = BORDER
    ws.cell(row=total_row2, column=1).alignment = ALIGN_CENTER

    # Section 3: 전체 요약 KPI
    r = total_row2 + 3
    ws.cell(row=r, column=1, value='3️⃣ 전체 요약').font = SECTION_FONT
    r += 1

    summary = [
        ('전체 문제 신고 건수', f'=COUNTA({sheet_problem_name}!A2:A51)'),
        ('  └ 🔴 일이 안 돼요', f'=COUNTIF({sheet_problem_name}!F2:F51,"{PROBLEM_LEVELS[0]}")'),
        ('  └ 🟡 불편하지만 우회', f'=COUNTIF({sheet_problem_name}!F2:F51,"{PROBLEM_LEVELS[1]}")'),
        ('  └ 🟢 사소', f'=COUNTIF({sheet_problem_name}!F2:F51,"{PROBLEM_LEVELS[2]}")'),
        ('전체 개선 제안 건수', f'=COUNTA({sheet_suggest_name}!A2:A51)'),
        ('  └ ⭐⭐⭐ 매일 도움', f'=COUNTIF({sheet_suggest_name}!F2:F51,"{SUGGESTION_PRIORITY[0]}")'),
        ('  └ ⭐⭐ 가끔 도움', f'=COUNTIF({sheet_suggest_name}!F2:F51,"{SUGGESTION_PRIORITY[1]}")'),
        ('  └ ⭐ 있으면 좋음', f'=COUNTIF({sheet_suggest_name}!F2:F51,"{SUGGESTION_PRIORITY[2]}")'),
        ('테스터 수 (문제 신고 기준)', f'=SUMPRODUCT((COUNTIF({sheet_problem_name}!A2:A51,{sheet_problem_name}!A2:A51)<>0)/COUNTIF({sheet_problem_name}!A2:A51,{sheet_problem_name}!A2:A51&""))'),
    ]
    for label, formula in summary:
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = BODY_FONT
        c1.alignment = Alignment(horizontal='left', vertical='center')
        c1.border = BORDER
        c2 = ws.cell(row=r, column=2, value=formula)
        c2.font = Font(name=FONT_NAME, size=12, bold=True)
        c2.alignment = ALIGN_CENTER
        c2.border = BORDER
        ws.row_dimensions[r].height = 22
        r += 1

    # Section 4: 다음 작업
    r += 2
    ws.cell(row=r, column=1, value='4️⃣ 관리자 매핑 가이드').font = SECTION_FONT
    r += 1
    mapping_note = (
        '본 사무직 폼은 직원이 자유롭게 입력합니다. UAT 관리자는 매일 다음을 수행:\n'
        '  1) 새로 추가된 문제·제안 검토\n'
        '  2) 개발자용 UAT_피드백_결과_v2.xlsx 피드백 로그에 FB-XXX 이슈로 옮겨 적기\n'
        '     - 화면 → 정확한 모듈 매핑\n'
        '     - 🔴/🟡/🟢 → Critical/Major/Minor 매핑 (필요 시 격상·격하)\n'
        '     - ⭐ → 개선 항목으로 분류\n'
        '  3) 중복 건은 사무직 폼 비고에 "FB-XXX와 동일" 표기 후 통합'
    )
    cell = ws.cell(row=r, column=1, value=mapping_note)
    cell.font = BODY_FONT
    cell.fill = CALLOUT_FILL
    cell.alignment = ALIGN_TOP_WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.row_dimensions[r].height = 140

    ws.sheet_view.zoomScale = 100


# ============================================================
# Main
# ============================================================
def main():
    wb = Workbook()
    wb.remove(wb.active)  # default sheet

    # Sheets in display order
    sheet_guide(wb)
    sheet_problem(wb)
    sheet_suggestion(wb)
    sheet_admin(wb)

    # Set first sheet active
    wb.active = 0

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    size_kb = OUT_PATH.stat().st_size / 1024
    print(f'✓ Wrote {OUT_PATH.relative_to(REPO_ROOT)} ({size_kb:.1f} KB)')
    print(f'  Sheets: {len(wb.sheetnames)}')
    for sn in wb.sheetnames:
        print(f'    - {sn}')


if __name__ == '__main__':
    main()
